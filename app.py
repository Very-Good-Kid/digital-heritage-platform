from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect

# 自定义CSRF保护类，支持通过属性豁免
class CustomCSRFProtect(CSRFProtect):
    def protect(self):
        if request.method not in ("GET", "HEAD", "OPTIONS", "TRACE"):
            # 检查是否有豁免标记
            if getattr(request, '_csrf_exempt', False):
                return

            # 检查视图函数是否有豁免属性
            view_func = self._get_view()
            if view_func and getattr(view_func, 'csrf_exempt', False):
                return

        return super().protect()

    def _get_view(self):
        """获取当前请求的视图函数"""
        endpoint = request.endpoint
        if endpoint and hasattr(self, 'app') and self.app:
            return self.app.view_functions.get(endpoint)
        return None
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, timezone
import os
from models import db, User, DigitalAsset, DigitalWill, PlatformPolicy, Story, FAQ
from utils.encryption import encrypt_data, decrypt_data
from utils.pdf_generator import generate_will_pdf
from config import config

# 注册后台管理蓝图
from admin import admin_bp

# 辅助函数
def get_china_time():
    """获取中国时区的当前时间"""
    utc_now = datetime.now(timezone.utc)
    china_tz = timezone(timedelta(hours=8))
    return utc_now.astimezone(china_tz)

# 初始化Flask应用
app = Flask(__name__)
app.config.from_object(config['default'])

# 性能优化: 静态文件缓存
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 31536000  # 1年缓存

# 确保数据目录存在（在应用启动前）
data_dir = app.config.get('DATA_DIR', 'instance')
if not os.path.exists(data_dir):
    try:
        os.makedirs(data_dir, exist_ok=True)
        print(f"Created data directory: {data_dir}")
    except Exception as e:
        print(f"Warning: Could not create data directory {data_dir}: {e}")

# 运行时字体安装（Render 免费版优化）
def setup_fonts_on_startup():
    """在应用启动时设置字体"""
    try:
        print("[INFO] Setting up fonts for PDF generation...")
        import subprocess
        import sys

        # 运行字体安装脚本
        script_path = os.path.join(os.path.dirname(__file__), 'install_fonts_runtime.py')
        if os.path.exists(script_path):
            result = subprocess.run(
                [sys.executable, script_path],
                capture_output=True,
                text=True,
                timeout=300
            )
            print(result.stdout)
            if result.stderr:
                print("[WARN]", result.stderr)
        else:
            print("[WARN] Font installation script not found, skipping...")

    except Exception as e:
        print(f"[WARN] Font setup failed: {e}")
        print("[INFO] Application will continue with default fonts")

# 初始化CSRF保护（使用自定义类支持属性豁免）
csrf = CustomCSRFProtect(app)



# 初始化扩展
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '请先登录访问此页面'

# 处理未登录请求 - API返回JSON，页面返回重定向
@login_manager.unauthorized_handler
def unauthorized():
    """处理未登录请求"""
    if request.path.startswith('/admin/api/'):
        return jsonify({'success': False, 'message': '请先登录'}), 401
    flash('请先登录访问此页面', 'warning')
    return redirect(url_for('login'))

# 注册蓝图
app.register_blueprint(admin_bp)

# 创建必要的目录
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('static/uploads', exist_ok=True)
os.makedirs('temp_pdfs', exist_ok=True)

# 自动初始化数据库（部署时自动执行）
def init_database_on_startup():
    """应用启动时自动初始化数据库表和数据"""
    with app.app_context():
        try:
            # 创建所有表
            db.create_all()
            print("[OK] Database tables created/verified")
            
            # 导入 PolicyDetail 模型
            from models import PolicyDetail
            
            # 检查是否需要初始化政策数据
            if PlatformPolicy.query.count() == 0:
                print("[INFO] Initializing platform policies...")
                init_default_policies()
            
            # 检查是否需要初始化政策详情数据
            if PolicyDetail.query.count() == 0:
                print("[INFO] Initializing policy details...")
                init_default_policy_details()
                
                # 迁移旧的资产分类
                migrate_asset_categories()

            # 检查是否需要初始化FAQ数据
            if FAQ.query.count() == 0:
                print("[INFO] Initializing FAQ data...")
                init_default_faqs()

            print("[OK] Database initialization complete")
        except Exception as e:
            print(f"[ERROR] Database initialization failed: {e}")

def init_default_policies():
    """初始化默认平台政策数据"""
    from models import PolicyDetail
    
    policies_data = [
        {
            'platform_name': 'QQ',
            'policy_content': '账号本身不可被继承；账号内财产可被继承',
            'attitude': '明确禁止',
            'inherit_possibility': '低',
            'customer_service': '综合热线：4006-700-700（9:00-22:00）；在线联系：腾讯客服公众号/小程序、网站kf.qq.com；本地备用：0755-83765566',
            'legal_basis': '《QQ号码规则》规定QQ号码所有权属于腾讯，使用权仅属于初始申请注册人，明确禁止继承。',
            'risk_warning': 'QQ账号长期未登录可能被回收，继承人可能永久失去访问渠道。'
        },
        {
            'platform_name': '微信',
            'policy_content': '账号本身不可被继承；账号内财产可被继承',
            'attitude': '明确禁止',
            'inherit_possibility': '低',
            'customer_service': '客服热线：95017；在线联系：微信APP内「我-设置-帮助与反馈」、微信/QQ端「腾讯客服」小程序',
            'legal_basis': '《腾讯微信软件许可及服务协议》规定微信账号所有权归腾讯，使用权仅属于初始申请注册人，明确禁止继承。',
            'risk_warning': '微信零钱余额需凭公证遗嘱、继承权公证书等向财付通公司申请提取。'
        },
        {
            'platform_name': '抖音',
            'policy_content': '账号本身原则上不可继承，但司法实践已有突破；账号内财产可被继承；逝者个人信息可依法复制/下载/转移',
            'attitude': '有限支持',
            'inherit_possibility': '中',
            'customer_service': '客服热线：95152；在线联系：抖音APP内「我-≡-我的客服」、微信/QQ端搜索抖音公众号；官方邮箱：feedback@douyin.com',
            'legal_basis': '《抖音隐私政策》3.5条规定逝者近亲属可依法行使个人信息相关权利，账号可设为纪念账号。',
            'risk_warning': '逝者近亲属需提交身份证明、死亡证明、亲属关系证明等材料完成核验。'
        }
    ]
    
    for data in policies_data:
        policy = PlatformPolicy(**data)
        db.session.add(policy)
    
    db.session.commit()
    print(f"[OK] Added {len(policies_data)} platform policies")


def migrate_asset_categories():
    """迁移旧的资产分类到新分类"""
    category_mapping = {
        '社交': '社交媒体',
        '金融': '电子邮箱',
        '记忆': '云存储与数字内容',
        '虚拟财产': '虚拟资产与数字货币'
    }

    updated_count = 0
    for old_cat, new_cat in category_mapping.items():
        assets = DigitalAsset.query.filter_by(category=old_cat).all()
        if assets:
            print(f"[INFO] Migrating {old_cat} -> {new_cat}: {len(assets)} assets")
            for asset in assets:
                asset.category = new_cat
            db.session.commit()
            updated_count += len(assets)

    if updated_count > 0:
        print(f"[OK] Migrated {updated_count} asset categories")

def init_default_policy_details():
    """初始化默认政策条款详情数据"""
    from models import PolicyDetail
    
    policy_details_data = {
        '微信': [
            {'policy_title': '《微信支付用户服务协议》第9.3条', 'policy_text': '"如你需要终止使用本服务时，你可以按照微信客户端的页面提示申请注销微信支付账户。……你的微信支付账户不得存在未处理完的交易或其他未了结的权利义务。"', 'legal_interpretation': '条款明确用户主动注销微信支付账户的前提。在继承场景中继承人要处理被继承人的微信支付账户，需先结清账户下所有未完成的交易。', 'display_order': 1},
            {'policy_title': '《腾讯服务协议》第3.1条', 'policy_text': '"腾讯服务账号的所有权归腾讯公司所有，用户完成申请注册手续后，仅获得腾讯服务账号的使用权，且该使用权仅属于初始申请注册人。"', 'legal_interpretation': '条款确立了"账号所有权归腾讯，使用权仅归初始注册人"的核心逻辑，从合同层面彻底排除了通过继承获得账号使用权的可能性。', 'display_order': 2},
            {'policy_title': '《微信支付用户服务协议》第1.6条', 'policy_text': '"你已知晓，\'零钱\'所记录的资金余额不同于你本人的银行存款，其实质为你委托财付通公司保管的、所有权归属于你的预付价值。"', 'legal_interpretation': '微信零钱余额是用户对财付通享有的债权，根据《民法典》继承规定，债权属于可继承的遗产范围。', 'display_order': 3},
            {'policy_title': '《腾讯微信软件许可及服务协议》7.1.2条', 'policy_text': '"微信账号的所有权归腾讯公司所有，用户完成申请注册手续后，仅获得微信账号的使用权，且该使用权仅属于初始申请注册人。"', 'legal_interpretation': '账号使用权与初始注册人的人身身份紧密绑定，具有高度的人身专属性。', 'display_order': 4},
            {'policy_title': '《腾讯微信软件许可及服务协议》7.1.5条', 'policy_text': '"针对基于微信账号创建的\'功能账号\'，同样规定非初始申请注册人不得通过受赠、继承、承租、受让或者其他任何方式使用该账号。"', 'legal_interpretation': '条款将"继承"明确列入禁止情形，从合同约定层面切断了继承人通过继承取得功能账号使用权的路径。', 'display_order': 5}
        ],
        'QQ': [
            {'policy_title': '《QQ号码规则》第二条：QQ号码的性质', 'policy_text': '"QQ号码是腾讯按照本规则授权注册用户用于登录、使用腾讯的软件或服务的数字标识，其所有权属于腾讯。"', 'legal_interpretation': 'QQ号码的所有权归属于腾讯公司，用户仅获得使用权，继承人无法主张对QQ号码本身的所有权。', 'display_order': 1},
            {'policy_title': '《QQ号码规则》第五条/第八条', 'policy_text': '"QQ号码使用权仅属于初始申请注册人。非初始申请注册人不得通过受赠、继承、承租、受让或者其他任何方式使用QQ号码。"', 'legal_interpretation': '条款明确禁止继承、赠与、借用、租用、转让、售卖等多种流转方式。', 'display_order': 2},
            {'policy_title': '《QQ软件许可及服务协议》第3条', 'policy_text': '若您注册的QQ号码长期没有登录或使用，腾讯有权将QQ号码进行回收处理。', 'legal_interpretation': '被继承人去世后，其QQ账号必然会进入长期未使用的状态，腾讯可依法回收账号。', 'display_order': 3},
            {'policy_title': '《QQ软件许可及服务协议》第8.1.5条、第8.3.3条', 'policy_text': '"腾讯将会尽其商业上的合理努力保障您在本服务中的数据存储安全，但是，腾讯并不能就此提供完全保证。"', 'legal_interpretation': '即便继承人能证明自身继承权，也可能因账号长期未使用被平台终止服务，导致数据被删除。', 'display_order': 4},
            {'policy_title': '《QQ软件许可及服务协议》第7.6条', 'policy_text': '"腾讯不会将您的个人信息转移或披露给任何第三方，除非相关法律法规或司法机关、行政机关要求。"', 'legal_interpretation': '继承人必须通过诉讼等司法程序取得法院调查令、判决书等法律文件，腾讯才有合同依据配合披露相关信息。', 'display_order': 5}
        ],
        '抖音': [
            {'policy_title': '《抖音隐私政策》3.3 复制、转移您的个人信息', 'policy_text': '如果您需要复制或下载我们收集存储的个人信息，您可以通过【我】-【☰】-【设置】-【个人信息管理】申请个人信息的下载。', 'legal_interpretation': '本条款的法定依据为《个人信息保护法》第四十五条规定的个人信息可携带权，逝者近亲属可依法行使该权利。', 'display_order': 1},
            {'policy_title': '《抖音隐私政策》3.4 账号注销', 'policy_text': '您可以在【我】-【☰】-【设置】-【账号与安全】-【注销账号】，或进入【抖音安全中心】进行账号注销。', 'legal_interpretation': '逝者近亲属若需注销逝者账号，需先完成亲属身份、逝者死亡证明等材料核验。', 'display_order': 2},
            {'policy_title': '《抖音隐私政策》3.5 逝者的个人信息保护', 'policy_text': '"如抖音用户不幸去世，其近亲属为了自身的合法、正当利益，可通过抖音手机客户端中的【我】-【☰】-【我的客服】与我们取得联系。"', 'legal_interpretation': '条款契合《个人信息保护法》第四十九条，逝者近亲属可对逝者账号及个人信息行使冻结、删除等权利，账号可设为纪念账号。', 'display_order': 3}
        ]
    }
    
    count = 0
    for platform_name, details in policy_details_data.items():
        platform = PlatformPolicy.query.filter_by(platform_name=platform_name).first()
        if not platform:
            continue
        
        for detail_data in details:
            detail = PolicyDetail(
                platform_policy_id=platform.id,
                policy_title=detail_data['policy_title'],
                policy_text=detail_data['policy_text'],
                legal_interpretation=detail_data['legal_interpretation'],
                display_order=detail_data['display_order']
            )
            db.session.add(detail)
            count += 1
    
    db.session.commit()
    print(f"[OK] Added {count} policy details")

def init_default_faqs():
    """初始化默认FAQ数据"""
    faqs_data = [
        FAQ(
            question='数字遗产包括哪些内容？',
            answer='数字遗产包括但不限于：社交媒体账号（微信、QQ、抖音等）、电子邮箱、云存储文件、虚拟货币、游戏账号、在线支付账户、博客文章、个人网站、数字相册、音视频文件等。',
            category='概念与价值',
            order=1
        ),
        FAQ(
            question='什么是数字遗嘱？',
            answer='数字遗嘱是指用户在生前制定的关于其数字资产如何处理的书面文件，包括账户信息、密码、处理方式等内容的详细说明。它可以指导继承人如何处理您的数字资产，避免账户丢失或数据永久消失。',
            category='概念与价值',
            order=2
        ),
        FAQ(
            question='为什么需要规划数字遗产？',
            answer='1. 避免重要数据永久丢失；2. 保护隐私和个人信息；3. 确保资产（如虚拟货币）不被浪费；4. 减轻家人的心理负担；5. 让数字记忆得以传承；6. 避免平台账户被自动删除。',
            category='概念与价值',
            order=3
        ),
        FAQ(
            question='数字资产的价值如何评估？',
            answer='数字资产的价值包括：经济价值（虚拟货币、游戏装备、付费内容等）、情感价值（照片、视频、聊天记录等）、实用价值（付费软件、云存储空间等）。建议定期整理和评估您的数字资产。',
            category='概念与价值',
            order=4
        ),
        FAQ(
            question='如何保护我的数字遗产？',
            answer='1. 定期备份重要数据到本地或云端；2. 使用密码管理器安全存储密码；3. 创建数字遗嘱并定期更新；4. 告知家人重要账户信息；5. 了解各平台的继承政策；6. 启用双重认证；7. 定期检查账户安全设置。',
            category='安全与管理',
            order=1
        ),
        FAQ(
            question='密码安全应该注意什么？',
            answer='1. 使用强密码（大小写字母、数字、特殊符号组合，至少12位）；2. 不要重复使用密码；3. 定期更换密码（每3-6个月）；4. 启用两步验证；5. 使用密码管理器（如LastPass、1Password）；6. 不要在公共场所输入密码；7. 警惕钓鱼网站。',
            category='安全与管理',
            order=2
        ),
        FAQ(
            question='如何选择密码管理器？',
            answer='选择密码管理器时考虑：1. 安全性（是否使用AES-256加密）；2. 跨平台支持（手机、电脑、浏览器）；3. 云同步功能；4. 价格（免费版功能）；5. 用户评价和口碑。推荐工具：LastPass、1Password、Bitwarden、KeePass等。',
            category='安全与管理',
            order=3
        ),
        FAQ(
            question='数字遗嘱有法律效力吗？',
            answer='数字遗嘱在我国法律体系中尚未明确认定，但可以作为表达意愿的重要依据。根据《民法典》，遗嘱可以采用多种形式，包括打印、录音录像等。数字遗嘱如果能证明是本人真实意愿，可能被参考。建议配合传统遗嘱使用，并咨询专业律师。',
            category='法律与继承',
            order=1
        ),
        FAQ(
            question='继承人的法律权利是什么？',
            answer='根据《民法典》，继承人有权继承被继承人的合法财产。但对于数字财产，法律界定尚不明确：1. 虚拟货币（比特币等）通常被视为财产，可以继承；2. 社交媒体账号（微信、QQ等）通常被视为使用权，不可继承；3. 游戏账号和虚拟道具的继承权取决于平台政策；4. 云存储文件可以继承，但需要密码或法律证明。',
            category='法律与继承',
            order=2
        ),
        FAQ(
            question='如何证明数字资产的所有权？',
            answer='证明数字资产所有权需要：1. 账户注册信息和登录记录；2. 交易记录和支付凭证（如购买虚拟货币的记录）；3. 电子邮件或聊天记录证明使用情况；4. 平台开具的资产证明；5. 公证处的公证文件；6. 银行流水证明充值记录。建议保留所有相关凭证。',
            category='法律与继承',
            order=3
        ),
        FAQ(
            question='如果平台拒绝继承怎么办？',
            answer='1. 查阅平台服务协议，了解具体条款；2. 准备完整的法律文件（死亡证明、继承公证书等）；3. 联系平台客服，说明情况；4. 寻求法律援助，向法院提起诉讼；5. 向消费者协会投诉；6. 通过媒体曝光引起关注。注意：不同平台处理方式不同，需要具体情况具体分析。',
            category='法律与继承',
            order=4
        ),
        FAQ(
            question='社交账号能否被作为遗产继承？',
            answer='不能。社交账号具有强烈的人身依附性，其使用权基于用户与平台的合同关系，不属于《民法典》第1122条规定的"遗产"范畴。但如今司法实践中有账号本身不能继承，但账号内的数据内容可以继承的判决倾向。法条依据：《民法典》第1122条："遗产是自然人死亡时遗留的个人合法财产。依照法律规定或者根据其性质不得继承的遗产，不得继承。"',
            category='法律与继承',
            order=5
        ),
        FAQ(
            question='游戏装备、虚拟货币等数字财产如何继承？',
            answer='可以继承，但是受平台用户协议限制。虚拟财产的继承需要平台配合，而平台常以"保护账号安全"为由拒绝。实务中游戏账号的充值余额、已购买未使用消耗的虚拟道具等可以明确估值的财产往往被认为可以继承，而游戏中的成就或其他有关人身性的财产被认定为不可继承。法条依据：《民法典》第127条："法律对数据、网络虚拟财产的保护有规定的，依照其规定。"确认虚拟财产受法律保护；第1122条将其纳入遗产范围。',
            category='法律与继承',
            order=6
        ),
        FAQ(
            question='电子邮件、聊天记录等内容是否能被继承人查阅？',
            answer='针对这类隐私性较强的财产应当区分对待。涉及到财产的邮件等应该能够依法被查阅（如银行账单、合同等）涉及到被继承人个人隐私的聊天记录与邮件内容则不能被查阅，除非被继承人同意。法条依据：《个人信息保护法》第49条："自然人死亡的，其近亲属为了自身的合法、正当利益，可以对死者的相关个人信息行使本章规定的查阅、复制、更正、删除等权利；死者生前另有安排的除外。"',
            category='伦理与隐私',
            order=1
        ),
        FAQ(
            question='如何保护逝者的隐私？',
            answer='1. 尊重逝者生前意愿；2. 不随意公开逝者的私人信息；3. 谨慎处理敏感内容；4. 遵守相关法律法规；5. 与家人协商处理方式；6. 必要时寻求专业法律建议。保护逝者隐私是对逝者的尊重，也是法律要求。',
            category='伦理与隐私',
            order=2
        ),
        FAQ(
            question='数字遗产继承中的伦理问题有哪些？',
            answer='1. 隐私与知情权的平衡；2. 逝者意愿与继承人利益的冲突；3. 数字身份的处理方式；4. 社交媒体纪念账号的管理；5. 敏感内容的处理；6. 家人之间的协商机制。建议在生前就这些问题做出明确安排。',
            category='伦理与隐私',
            order=3
        )
    ]

    for faq in faqs_data:
        db.session.add(faq)

    db.session.commit()
    print(f"[OK] Added {len(faqs_data)} FAQs")

# 使用 before_request 替代模块级别调用
# 这样可以确保应用完全初始化后再执行数据库初始化
_db_initialized = False

@app.before_request
def initialize_database():
    """在第一个请求前初始化数据库（仅执行一次）"""
    global _db_initialized
    if not _db_initialized:
        _db_initialized = True
        try:
            db.create_all()
            print("[OK] Database tables created/verified")
            
            from models import PolicyDetail
            
            if PlatformPolicy.query.count() == 0:
                print("[INFO] Initializing platform policies...")
                init_default_policies()
            
            if PolicyDetail.query.count() == 0:
                print("[INFO] Initializing policy details...")
                init_default_policy_details()

            migrate_asset_categories()

            if FAQ.query.count() == 0:
                print("[INFO] Initializing FAQ data...")
                init_default_faqs()

            print("[OK] Database initialization complete")
        except Exception as e:
            print(f"[ERROR] Database initialization failed: {e}")

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# 设置响应编码
@app.after_request
def after_request(response):
    """确保响应使用UTF-8编码"""
    # 只修改HTML响应，不影响JSON等API响应
    content_type = response.headers.get('Content-Type', '')
    if 'text/html' in content_type or not content_type:
        # 如果没有设置Content-Type或为HTML，则设置为UTF-8
        if '; charset' not in content_type:
            if content_type:
                response.headers['Content-Type'] = f"{content_type}; charset=utf-8"
            else:
                response.headers['Content-Type'] = 'text/html; charset=utf-8'
    return response

# 数据库连接健康检查中间件
@app.before_request
def check_db_connection():
    """在每个请求前检查数据库连接 - 优化版本"""
    # 只在特定路由检查,避免每个请求都检查
    if request.endpoint and request.endpoint.startswith('admin'):
        try:
            # 执行一个简单的查询来检查连接
            db.session.execute(db.text('SELECT 1'))
        except Exception as e:
            # 如果连接失败,尝试重连
            print(f"[WARN] Database connection lost, attempting to reconnect: {e}")
            try:
                db.session.remove()
                db.session.execute(db.text('SELECT 1'))
                print("[INFO] Database reconnection successful")
            except Exception as e2:
                print(f"[ERROR] Database reconnection failed: {e2}")
                # 如果重连失败,返回错误页面
                return "数据库连接失败,请稍后重试", 503

# 路由：处理 Chrome DevTools 请求（避免 404 报错）
@app.route('/.well-known/appspecific/com.chrome.devtools.json')
def chrome_devtools():
    """处理 Chrome DevTools 的配置请求"""
    return '{}', 200, {'Content-Type': 'application/json'}

# 路由：favicon
@app.route('/favicon.ico')
def favicon():
    """网站图标 - 使用简单的SVG图标"""
    from flask import Response
    # 创建一个简单的SVG图标
    svg_icon = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32">
        <rect width="32" height="32" fill="#667eea" rx="6"/>
        <text x="16" y="22" font-size="18" fill="white" text-anchor="middle" font-weight="bold">D</text>
    </svg>"""
    return Response(svg_icon, mimetype='image/svg+xml')

# 路由：首页
@app.route('/')
def index():
    """首页"""
    return render_template('index.html')

# 路由：用户认证
@app.route('/register', methods=['GET', 'POST'])
def register():
    """用户注册"""
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if not all([username, email, password, confirm_password]):
            flash('请填写所有字段', 'error')
            return redirect(url_for('register'))

        if password != confirm_password:
            flash('两次输入的密码不一致', 'error')
            return redirect(url_for('register'))

        if User.query.filter_by(username=username).first():
            flash('用户名已存在', 'error')
            return redirect(url_for('register'))

        if User.query.filter_by(email=email).first():
            flash('邮箱已被注册', 'error')
            return redirect(url_for('register'))

        user = User(
            username=username,
            email=email,
            password_hash=generate_password_hash(password)
        )

        try:
            db.session.add(user)
            db.session.commit()
            flash('注册成功，请登录', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash('注册失败，请重试', 'error')
            return redirect(url_for('register'))

    return render_template('auth/register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """用户登录"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = request.form.get('remember', False)

        user = User.query.filter_by(username=username).first()

        if user and check_password_hash(user.password_hash, password):
            login_user(user, remember=remember)
            next_page = request.args.get('next')

            # 如果指定了next页面，跳转到next页面
            if next_page:
                return redirect(next_page)

            # 管理员跳转到后台，普通用户跳转到用户仪表盘
            if user.is_admin:
                return redirect(url_for('admin.dashboard'))
            else:
                return redirect(url_for('dashboard'))
        else:
            flash('用户名或密码错误', 'error')

    return render_template('auth/login.html')

@app.route('/logout')
@login_required
def logout():
    """用户登出"""
    logout_user()
    flash('已成功登出', 'success')
    return redirect(url_for('login'))

# 路由：用户仪表盘
@app.route('/dashboard')
@login_required
def dashboard():
    """用户仪表盘"""
    # 应用多租户数据隔离
    assets = DigitalAsset.query
    wills = DigitalWill.query

    # 管理员可以看所有数据，普通用户只能看自己的数据
    if not current_user.is_admin:
        assets = assets.filter_by(user_id=current_user.id)
        wills = wills.filter_by(user_id=current_user.id)

    assets = assets.order_by(DigitalAsset.created_at.desc()).all()
    wills = wills.order_by(DigitalWill.created_at.desc()).all()
    return render_template('dashboard/index.html', assets=assets, wills=wills)

# 路由：数字资产清单
@app.route('/assets', methods=['GET', 'POST'])
@login_required
def assets():
    """数字资产清单"""
    if request.method == 'POST':
        platform_name = request.form.get('platform_name')
        account = request.form.get('account')
        password = request.form.get('password')
        category = request.form.get('category')
        notes = request.form.get('notes')

        encrypted_password = encrypt_data(password) if password else None

        asset = DigitalAsset(
            user_id=current_user.id,
            platform_name=platform_name,
            account=account,
            encrypted_password=encrypted_password,
            category=category,
            notes=notes
        )

        try:
            db.session.add(asset)
            db.session.commit()
            flash('数字资产添加成功', 'success')
        except Exception as e:
            db.session.rollback()
            flash('添加失败，请重试', 'error')

        return redirect(url_for('assets'))

    # 应用多租户数据隔离
    assets_query = DigitalAsset.query
    if not current_user.is_admin:
        assets_query = assets_query.filter_by(user_id=current_user.id)

    assets = assets_query.order_by(DigitalAsset.created_at.desc()).all()
    
    # 获取用户的遗嘱列表
    wills_query = DigitalWill.query
    if not current_user.is_admin:
        wills_query = wills_query.filter_by(user_id=current_user.id)
    wills = wills_query.order_by(DigitalWill.created_at.desc()).all()
    
    categories = ['社交媒体', '电子邮箱', '云存储与数字内容', '虚拟资产与数字货币', '其他数字资产']
    return render_template('assets/index.html', assets=assets, categories=categories, wills=wills)

@app.route('/assets/download-template/<format>')
@login_required
def download_template(format):
    """下载数字资产模板"""
    if format == 'excel':
        return generate_excel_template()
    elif format == 'pdf':
        return generate_pdf_template()
    else:
        flash('不支持的模板格式', 'error')
        return redirect(url_for('assets'))

def generate_excel_template():
    """生成Excel模板"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "数字资产清单"

        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40

        # 定义样式
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头
        headers = ['平台名称', '账号', '密码（可选）', '分类', '备注说明']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 添加示例数据
        sample_data = [
            ['微信', '13800138000', '', '社交媒体', '主要社交账号'],
            ['QQ', '123456789', '', '社交媒体', '工作用QQ'],
            ['支付宝', '13800138000', 'password123', '虚拟资产与数字货币', '主要支付账户'],
            ['百度网盘', 'user@example.com', 'pwd123', '云存储与数字内容', '存储重要文件'],
            ['王者荣耀', 'game_user', 'gamepass', '虚拟资产与数字货币', '游戏账号']
        ]

        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

        # 添加说明
        ws.append([])
        ws.append(['填写说明：'])
        ws.append(['1. 平台名称：填写数字资产所属平台（如微信、QQ、抖音等）'])
        ws.append(['2. 账号：填写您的账号信息（手机号、邮箱、用户名等）'])
        ws.append(['3. 密码：选填，密码将加密存储'])
        ws.append(['4. 分类：选择资产分类（社交媒体、电子邮箱、云存储与数字内容、虚拟资产与数字货币、其他数字资产）'])
        ws.append(['5. 备注：填写任何有用的补充信息'])

        # 保存到内存
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='数字资产清单模板.xlsx'
        )
    except ImportError:
        flash('Excel模板生成功能需要安装openpyxl库', 'error')
        return redirect(url_for('assets'))
    except Exception as e:
        flash(f'模板生成失败：{str(e)}', 'error')
        return redirect(url_for('assets'))

def generate_pdf_template():
    """生成PDF模板"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from io import BytesIO

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)

        # 创建样式
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            spaceAfter=12,
            spaceBefore=15
        )

        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=10,
            spaceAfter=10,
            leading=14
        )

        story = []

        # 标题
        story.append(Paragraph("数字资产清单模板", title_style))
        story.append(Spacer(1, 0.2 * inch))

        # 说明
        story.append(Paragraph("<b>填写说明：</b>", heading_style))
        story.append(Paragraph(
            "请按照以下格式填写您的数字资产信息。填写完成后，可以将此表格作为参考，"
            "在网站中添加您的数字资产。",
            body_style
        ))
        story.append(Spacer(1, 0.3 * inch))

        # 表格
        table_data = [
            ['平台名称', '账号', '密码（可选）', '分类', '备注说明'],
            ['微信', '13800138000', '', '社交媒体', '主要社交账号'],
            ['QQ', '123456789', '', '社交媒体', '工作用QQ'],
            ['支付宝', '13800138000', '', '虚拟资产与数字货币', '主要支付账户'],
            ['百度网盘', 'user@example.com', '', '云存储与数字内容', '存储重要文件'],
            ['王者荣耀', 'game_user', '', '虚拟资产与数字货币', '游戏账号'],
            ['', '', '', '', ''],
            ['', '', '', '', ''],
            ['', '', '', '', ''],
            ['', '', '', '', ''],
        ]

        table = Table(table_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 1*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))

        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

        # 详细说明
        story.append(Paragraph("<b>字段说明：</b>", heading_style))
        story.append(Paragraph(
            "<b>1. 平台名称：</b>填写数字资产所属平台（如微信、QQ、抖音、支付宝等）",
            body_style
        ))
        story.append(Paragraph(
            "<b>2. 账号：</b>填写您的账号信息（手机号、邮箱、用户名等）",
            body_style
        ))
        story.append(Paragraph(
            "<b>3. 密码：</b>选填，密码将加密存储，请妥善保管",
            body_style
        ))
        story.append(Paragraph(
            "<b>4. 分类：</b>选择资产分类（社交媒体、电子邮箱、云存储与数字内容、虚拟资产与数字货币、其他数字资产）",
            body_style
        ))
        story.append(Paragraph(
            "<b>5. 备注：</b>填写任何有用的补充信息",
            body_style
        ))
        story.append(Spacer(1, 0.3 * inch))

        # 分类说明
        story.append(Paragraph("<b>资产分类说明：</b>", heading_style))
        story.append(Paragraph(
            "<b>社交：</b>社交媒体账号（微信、QQ、微博、抖音等）",
            body_style
        ))
        story.append(Paragraph(
            "<b>金融：</b>金融账户（支付宝、微信支付、银行网银、投资理财等）",
            body_style
        ))
        story.append(Paragraph(
            "<b>记忆：</b>记忆存储（云存储、相册、文档、笔记等）",
            body_style
        ))
        story.append(Paragraph(
            "<b>虚拟财产：</b>虚拟资产（游戏账号、虚拟货币、其他虚拟物品）",
            body_style
        ))

        # 生成PDF
        doc.build(story)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name='数字资产清单模板.pdf'
        )
    except Exception as e:
        flash(f'PDF模板生成失败：{str(e)}', 'error')
        return redirect(url_for('assets'))

@app.route('/assets/<int:asset_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_asset(asset_id):
    """编辑数字资产"""
    asset = DigitalAsset.query.get_or_404(asset_id)

    if asset.user_id != current_user.id:
        flash('无权访问此资产', 'error')
        return redirect(url_for('assets'))

    if request.method == 'POST':
        asset.platform_name = request.form.get('platform_name')
        asset.account = request.form.get('account')
        password = request.form.get('password')
        asset.category = request.form.get('category')
        asset.notes = request.form.get('notes')

        if password:
            asset.encrypted_password = encrypt_data(password)

        asset.updated_at = get_china_time()

        try:
            db.session.commit()
            flash('资产更新成功', 'success')
            return redirect(url_for('assets'))
        except Exception as e:
            db.session.rollback()
            flash('更新失败，请重试', 'error')

    categories = ['社交媒体', '电子邮箱', '云存储与数字内容', '虚拟资产与数字货币', '其他数字资产']
    return render_template('assets/edit.html', asset=asset, categories=categories)

@app.route('/assets/<int:asset_id>/delete', methods=['POST'])
@login_required
def delete_asset(asset_id):
    """删除数字资产"""
    asset = DigitalAsset.query.get_or_404(asset_id)

    if asset.user_id != current_user.id:
        flash('无权访问此资产', 'error')
        return redirect(url_for('assets'))

    try:
        db.session.delete(asset)
        db.session.commit()
        flash('资产删除成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash('删除失败，请重试', 'error')

    return redirect(url_for('assets'))

@app.route('/assets/<int:asset_id>/decrypt', methods=['POST'])
@login_required
def decrypt_asset(asset_id):
    """解密资产密码"""
    asset = DigitalAsset.query.get_or_404(asset_id)

    if asset.user_id != current_user.id:
        flash('无权访问此资产', 'error')
        return redirect(url_for('assets'))

    if asset.encrypted_password:
        try:
            decrypted_password = decrypt_data(asset.encrypted_password)
            return jsonify({'success': True, 'password': decrypted_password})
        except Exception as e:
            return jsonify({'success': False, 'message': '解密失败'})
    else:
        return jsonify({'success': False, 'message': '无密码信息'})

# 路由：数字遗嘱
@app.route('/wills', methods=['GET', 'POST'])
@login_required
def wills():
    """数字遗嘱列表"""
    if request.method == 'POST':
        assets_data = request.form.get('assets_data', '{}')

        try:
            assets_data_json = __import__('json').loads(assets_data)
        except:
            assets_data_json = {}

        will = DigitalWill(
            user_id=current_user.id,
            title='数字资产继承意愿声明书',  # 使用固定标题
            description='',
            assets_data=assets_data_json,
            status='draft'
        )

        try:
            db.session.add(will)
            db.session.commit()
            flash('数字资产继承意愿声明书创建成功', 'success')
            return redirect(url_for('wills'))
        except Exception as e:
            db.session.rollback()
            flash('创建失败，请重试', 'error')

    # 应用多租户数据隔离
    wills_query = DigitalWill.query
    assets_query = DigitalAsset.query

    if not current_user.is_admin:
        wills_query = wills_query.filter_by(user_id=current_user.id)
        assets_query = assets_query.filter_by(user_id=current_user.id)

    wills = wills_query.order_by(DigitalWill.created_at.desc()).all()
    assets = assets_query.all()
    return render_template('wills/index.html', wills=wills, assets=assets)

@app.route('/wills/download-template/<format>')
@login_required
def download_will_template(format):
    """下载遗嘱模板"""
    if format == 'excel':
        return generate_will_excel_template()
    elif format == 'pdf':
        return generate_will_pdf_template()
    else:
        flash('不支持的模板格式', 'error')
        return redirect(url_for('wills'))

def generate_will_excel_template():
    """生成遗嘱Excel模板"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "数字遗产意愿声明"

        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40

        # 定义样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        header_font = Font(name='微软雅黑', size=12, bold=True)
        header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 标题
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = '数字遗产意愿声明模板'
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 基本信息
        ws.append([])
        ws.append(['基本信息'])
        ws.append(['遗嘱标题', '我的数字遗产处理意愿'])
        ws.append(['总体意愿说明', '请说明您对数字遗产的总体处理意愿'])
        ws.append(['指定继承人信息', '填写继承人的姓名、关系、联系方式等信息'])
        ws.append(['特别说明', '其他需要特别说明的事项'])

        # 数字资产处理表
        ws.append([])
        ws.append(['数字资产处理意愿'])
        headers = ['平台名称', '账号', '处理方式']
        ws.append(headers)

        # 设置表头样式
        for col in range(1, 4):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # 示例数据
        sample_data = [
            ['微信', '13800138000', '指定继承人'],
            ['QQ', '123456789', '转为纪念模式'],
            ['支付宝', '13800138000', '指定继承人'],
            ['百度网盘', 'user@example.com', '委托删除'],
        ]

        for row_data in sample_data:
            ws.append(row_data)
            for col in range(1, 4):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.border = border

        # 处理方式说明
        ws.append([])
        ws.append(['处理方式说明：'])
        ws.append(['1. 指定继承人：将数字资产转移给指定的继承人'])
        ws.append(['2. 转为纪念模式：保持账户活跃状态，但不进行登录操作'])
        ws.append(['3. 委托删除：授权平台或继承人删除账户'])
        ws.append(['4. 其他处理方式：根据具体情况自定义处理方式'])

        # 法律提示
        ws.append([])
        ws.append(['重要法律提示：'])
        ws.append(['本文件可作为您意愿的强烈表达，协助继承人沟通，'])
        ws.append(['但不替代正式公证遗嘱。建议咨询专业律师。'])

        # 保存到内存
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='数字遗产意愿声明模板.xlsx'
        )
    except ImportError:
        flash('Excel模板生成功能需要安装openpyxl库', 'error')
        return redirect(url_for('wills'))
    except Exception as e:
        flash(f'模板生成失败：{str(e)}', 'error')
        return redirect(url_for('wills'))

def generate_will_pdf_template():
    """生成遗嘱PDF模板"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from io import BytesIO

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)

        # 创建样式
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=20,
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            spaceAfter=12,
            spaceBefore=20
        )

        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=11,
            spaceAfter=12,
            alignment=TA_JUSTIFY,
            leading=16
        )

        warning_style = ParagraphStyle(
            'CustomWarning',
            parent=styles['BodyText'],
            fontName='Helvetica-Oblique',
            fontSize=9,
            textColor=colors.red,
            spaceAfter=20,
            alignment=TA_CENTER
        )

        story = []

        # 标题
        story.append(Paragraph("数字遗产意愿声明", title_style))
        story.append(Spacer(1, 0.3 * inch))

        # 基本信息说明
        story.append(Paragraph("<b>一、基本信息</b>", heading_style))
        story.append(Paragraph(
            "<b>遗嘱标题：</b>请填写遗嘱标题，例如：我的数字遗产处理意愿",
            body_style
        ))
        story.append(Paragraph(
            "<b>总体意愿说明：</b>请简要说明您对数字遗产的总体处理意愿",
            body_style
        ))
        story.append(Paragraph(
            "<b>指定继承人信息：</b>请填写继承人的姓名、关系、联系方式等详细信息",
            body_style
        ))
        story.append(Paragraph(
            "<b>特别说明：</b>请填写其他需要特别说明的事项",
            body_style
        ))
        story.append(Spacer(1, 0.3 * inch))

        # 数字资产处理表
        story.append(Paragraph("<b>二、数字资产处理意愿</b>", heading_style))
        story.append(Paragraph(
            "请为每个数字资产选择处理方式，并在下表中填写详细信息：",
            body_style
        ))

        table_data = [
            ['平台名称', '账号', '处理方式', '备注'],
            ['微信', '13800138000', '指定继承人', '主要社交账号'],
            ['QQ', '123456789', '转为纪念模式', '工作用QQ'],
            ['支付宝', '13800138000', '指定继承人', '主要支付账户'],
            ['百度网盘', 'user@example.com', '委托删除', '存储重要文件'],
            ['', '', '', ''],
            ['', '', '', ''],
            ['', '', '', ''],
        ]

        table = Table(table_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))

        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

        # 处理方式说明
        story.append(Paragraph("<b>三、处理方式说明</b>", heading_style))
        story.append(Paragraph(
            "<b>1. 指定继承人：</b>将数字资产转移给指定的继承人，需要提供继承人的联系信息。",
            body_style
        ))
        story.append(Paragraph(
            "<b>2. 转为纪念模式：</b>保持账户活跃状态，保留聊天记录和分享内容，但不进行登录操作。",
            body_style
        ))
        story.append(Paragraph(
            "<b>3. 委托删除：</b>授权平台或继承人删除账户，清除所有个人数据。",
            body_style
        ))
        story.append(Paragraph(
            "<b>4. 其他处理方式：</b>根据具体情况自定义处理方式，需要详细说明处理流程。",
            body_style
        ))
        story.append(Spacer(1, 0.3 * inch))

        # 平台政策参考
        story.append(Paragraph("<b>四、平台政策参考</b>", heading_style))
        story.append(Paragraph(
            "请注意，各数字平台对账户继承有不同的政策规定。在执行本声明时，"
            "建议继承人提前了解相关平台的具体政策，并准备必要的法律文件。",
            body_style
        ))
        story.append(Paragraph(
            "<b>微信：</b>账户所有权归腾讯公司所有，继承需要提供相关证明材料。",
            body_style
        ))
        story.append(Paragraph(
            "<b>QQ：</b>可以申请继承，需要提供死亡证明、亲属关系证明等材料。",
            body_style
        ))
        story.append(Paragraph(
            "<b>抖音：</b>继承政策相对严格，需要提供法律文件和身份证明。",
            body_style
        ))
        story.append(Spacer(1, 0.3 * inch))

        # 法律提示
        story.append(Paragraph(
            "<b>五、重要法律提示</b>",
            heading_style
        ))
        story.append(Paragraph(
            "本文件可作为您意愿的强烈表达，协助继承人沟通，但不替代正式公证遗嘱。"
            "建议在制定此声明后，咨询专业律师，并考虑进行公证。",
            warning_style
        ))
        story.append(Spacer(1, 0.3 * inch))

        # 免责声明
        story.append(Paragraph(
            "<b>六、免责声明</b>",
            heading_style
        ))
        story.append(Paragraph(
            "本平台仅提供信息参考和工具服务，不构成法律建议。本声明的执行情况"
            "取决于各平台的具体政策、相关法律法规以及继承人的实际操作。"
            "因执行本声明而产生的任何法律纠纷，本平台不承担任何责任。",
            body_style
        ))

        # 生成PDF
        doc.build(story)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name='数字遗产意愿声明模板.pdf'
        )
    except Exception as e:
        flash(f'PDF模板生成失败：{str(e)}', 'error')
        return redirect(url_for('wills'))

@app.route('/wills/<int:will_id>/view', methods=['GET'])
@login_required
def view_will(will_id):
    """查看数字遗嘱"""
    will = DigitalWill.query.get_or_404(will_id)

    if will.user_id != current_user.id:
        flash('无权访问此遗嘱', 'error')
        return redirect(url_for('wills'))

    return render_template('wills/view.html', will=will)

@app.route('/wills/<int:will_id>/generate_pdf', methods=['GET'])
@login_required
def generate_pdf(will_id):
    """生成遗嘱PDF"""
    will = DigitalWill.query.get_or_404(will_id)

    if will.user_id != current_user.id:
        flash('无权访问此遗嘱', 'error')
        return redirect(url_for('wills'))

    try:
        print(f"Generating PDF for will {will_id}: {will.title}")
        pdf_path = generate_will_pdf(will)
        print(f"PDF generated at: {pdf_path}")

        if pdf_path and os.path.exists(pdf_path):
            return send_file(pdf_path, as_attachment=True, download_name=f'数字遗嘱_{will.title}.pdf')
        else:
            flash('PDF文件生成失败，请稍后重试', 'error')
            return redirect(url_for('view_will', will_id=will_id))
    except Exception as e:
        print(f"PDF generation error: {e}")
        import traceback
        traceback.print_exc()
        flash(f'PDF生成失败：{str(e)}', 'error')
        return redirect(url_for('view_will', will_id=will_id))

@app.route('/wills/<int:will_id>/delete', methods=['POST'])
@login_required
def delete_will(will_id):
    """删除数字遗嘱"""
    will = DigitalWill.query.get_or_404(will_id)

    if will.user_id != current_user.id:
        flash('无权访问此遗嘱', 'error')
        return redirect(url_for('wills'))

    try:
        db.session.delete(will)
        db.session.commit()
        flash('遗嘱删除成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash('删除失败，请重试', 'error')

    return redirect(url_for('wills'))


@app.route('/wills/<int:will_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_will(will_id):
    """编辑数字遗嘱"""
    will = DigitalWill.query.get_or_404(will_id)

    # 权限控制：只能编辑自己的遗嘱
    if will.user_id != current_user.id:
        flash('无权访问此遗嘱', 'error')
        return redirect(url_for('wills'))

    if request.method == 'POST':
        assets_data = request.form.get('assets_data', '{}')
        new_status = request.form.get('status', will.status)

        # 状态变更权限控制
        # 只有管理员可以设置confirmed或archived状态
        # 用户自己只能设置为draft状态
        if new_status != will.status:
            if not current_user.is_admin:
                if new_status != 'draft':
                    flash('只有管理员可以将遗嘱状态设置为已确认或已归档', 'error')
                    return redirect(url_for('edit_will', will_id=will_id))
            else:
                # 管理员可以更改任何状态
                will.status = new_status

        # 验证JSON格式
        try:
            if assets_data:
                import json
                assets_data_json = json.loads(assets_data)
            else:
                assets_data_json = {}
        except json.JSONDecodeError:
            assets_data_json = {}

        # 更新其他字段
        will.assets_data = assets_data_json
        will.updated_at = get_china_time()

        try:
            db.session.commit()
            flash('遗嘱更新成功', 'success')
            return redirect(url_for('wills'))
        except Exception as e:
            db.session.rollback()
            flash('更新失败，请重试', 'error')

    assets = DigitalAsset.query.filter_by(user_id=current_user.id).all()
    return render_template('wills/edit.html', will=will, assets=assets)



@app.route('/wills/<int:will_id>/status', methods=['POST'])
@login_required
def update_will_status(will_id):
    """更新遗嘱状态"""
    will = DigitalWill.query.get_or_404(will_id)

    # 权限控制
    if will.user_id != current_user.id and not current_user.is_admin:
        flash('无权访问此遗嘱', 'error')
        return redirect(url_for('wills'))

    new_status = request.json.get('status')

    # 状态值验证
    valid_statuses = ['draft', 'confirmed', 'archived']
    if new_status not in valid_statuses:
        return jsonify({'success': False, 'message': '无效的状态值'}), 400

    # 业务规则：
    # 1. 用户只能在"草稿"和"已确认"之间切换
    # 2. 用户不能将遗嘱改为"已归档"
    # 3. 管理员可以更改任何状态，无限制
    if not current_user.is_admin:
        # 非管理员用户
        # 允许在草稿和已确认之间切换
        if new_status in ['draft', 'confirmed']:
            if will.status != new_status:
                will.status = new_status
        else:
            # 用户尝试改为已归档，拒绝
            return jsonify({
                'success': False,
                'message': '只有管理员可以将遗嘱状态设置为已归档'
            }), 403
    else:
        # 管理员可以更改任何状态，无限制
        will.status = new_status

    will.updated_at = get_china_time()



    try:
        db.session.commit()
        return jsonify({
            'success': True,
            'message': '遗嘱状态更新成功',
            'new_status': will.status
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'}), 500

# 路由：平台政策矩阵
@app.route('/policies')
def policies():
    """平台政策矩阵"""
    policies = PlatformPolicy.query.order_by(PlatformPolicy.platform_name).all()
    return render_template('policies/index.html', policies=policies)

# 路由：身后继承（整合平台政策和继承导航）
@app.route('/inheritance')
def inheritance():
    """身后继承 - 整合平台政策矩阵和继承导航（静态版本）"""
    scenarios = [
        {'id': 'scenario1', 'name': '有遗嘱+有密码', 'description': '您拥有合法的数字遗嘱和账户密码'},
        {'id': 'scenario2', 'name': '有遗嘱+无密码', 'description': '您拥有合法的数字遗嘱但没有账户密码'},
        {'id': 'scenario3', 'name': '无遗嘱+无密码', 'description': '您没有数字遗嘱和账户密码'}
    ]
    return render_template('inheritance/index.html', scenarios=scenarios)

# 路由：继承导航
@app.route('/inheritance-guide', methods=['GET', 'POST'])
def inheritance_guide():
    """数字资产继承导航"""
    if request.method == 'POST':
        platform = request.form.get('platform')
        scenario = request.form.get('scenario')
        
        # 验证参数
        if not platform or not scenario:
            flash('请选择平台和继承情景', 'error')
            return redirect(url_for('inheritance_guide'))
        
        return redirect(url_for('inheritance_result', platform=platform, scenario=scenario))

    platforms = PlatformPolicy.query.filter(PlatformPolicy.platform_name.in_(['微信', 'QQ', '抖音'])).all()
    scenarios = [
        {'id': 'scenario1', 'name': '有遗嘱+有密码', 'description': '您拥有合法的数字遗嘱和账户密码'},
        {'id': 'scenario2', 'name': '有遗嘱+无密码', 'description': '您拥有合法的数字遗嘱但没有账户密码'},
        {'id': 'scenario3', 'name': '无遗嘱+无密码', 'description': '您没有数字遗嘱和账户密码'}
    ]
    return render_template('inheritance-guide/index.html', platforms=platforms, scenarios=scenarios)

@app.route('/inheritance-result')
def inheritance_result():
    """继承导航结果"""
    platform = request.args.get('platform')
    scenario = request.args.get('scenario')

    policy = PlatformPolicy.query.filter_by(platform_name=platform).first()
    if not policy:
        flash('未找到相关平台信息', 'error')
        return redirect(url_for('inheritance_guide'))

    steps = generate_inheritance_steps(platform, scenario, policy)
    return render_template('inheritance-guide/result.html',
                         platform=platform,
                         scenario=scenario,
                         policy=policy,
                         steps=steps)

def generate_inheritance_steps(platform, scenario, policy):
    """生成继承步骤"""
    steps = []

    # 平台专属信息配置
    platform_info = {
        'QQ': {
            'customer_service': '4006700700',
            'official_channels': ['腾讯客服官网(kf.qq.com)', '微信/QQ端「腾讯客服」小程序'],
            'inheritable_assets': 'QQ钱包余额、理财通资产、Q币、腾讯系游戏道具/装备、付费会员可转移权益、可变现原创数字内容等',
            'owner': '腾讯科技（深圳）有限公司'
        },
        '微信': {
            'customer_service': '95017',
            'official_channels': ['微信APP内「我-设置-帮助与反馈」', '微信/QQ端「腾讯客服」小程序'],
            'inheritable_assets': '微信支付余额、理财通资产、微信零钱通、视频号创作收益、公众号相关可变现资产、付费会员可转移权益等',
            'owner': '腾讯科技（深圳）有限公司'
        },
        '抖音': {
            'customer_service': '95152',
            'official_channels': ['抖音APP内「我-设置-反馈与帮助」', '抖音安全中心官方公众号'],
            'inheritable_assets': '抖币余额、直播带货收益、电商橱窗资产、创作收入、付费课程可转移权益、本地生活可核销权益等',
            'owner': '北京字节跳动网络技术有限公司'
        }
    }

    info = platform_info.get(platform, platform_info['QQ'])

    if scenario == 'scenario1':  # 有遗嘱+有密码
        # 微信平台专属步骤（根据PDF文件）
        if platform == '微信':
            steps = [
                {
                    'step': 1,
                    'title': '明确继承目标',
                    'description': '',
                    'materials_table': True,
                    'materials': [
                        {'name': '账号本身（登录权、使用权等）', 'detail': '平台用户协议明确禁止账号继承、转让等', 'extra': '不可继承'},
                        {'name': '账号内虚拟财产', 'detail': '准备基础证明文件，按平台及法律流程办理', 'extra': '可依法继承'}
                    ],
                    'templates': [],
                    'warning': '【场景合规提示】即使您知晓逝者微信账号的密码，也严禁利用密码进行任何未经授权或超出继承申请范围的操作，例如：擅自转移大额资金、删除聊天记录等原始数据、冒用身份进行社交或交易。此类行为可能违反平台协议、遗嘱本意及相关法律，可能导致您的继承申请被拒绝，甚至需要承担法律责任。'
                },
                {
                    'step': 2,
                    'title': '办理继承权公证',
                    'description': '材料齐全且无误后，前往公证处办理。（详见：《数字遗产继承公证办理流程说明》）',
                    'materials': [
                        {'name': '被继承人死亡证明', 'detail': '医院或派出所等部门出具'},
                        {'name': '继承关系证明', 'detail': '如户口本、结婚证等'},
                        {'name': '继承人身份证明', 'detail': '继承人的身份证、户口簿等有效证件'},
                        {'name': '财产凭证', 'detail': '如微信钱包余额截图、虚拟商品购买记录等，证明账号内虚拟财产的存在与价值'}
                    ],
                    'templates': ['数字遗产继承公证办理流程说明']
                },
                {
                    'step': 3,
                    'title': '材料整理与遗嘱效力核验',
                    'description': '提前备齐所有符合平台要求、具备法律效力的申请材料，核验遗嘱法定效力，梳理已知遗产明细，避免因材料不全、不合规被驳回。',
                    'materials': [
                        {'name': '合法有效遗嘱原件', 'detail': '符合《民法典》法定形式，纸质原件彩色扫描件；公证遗嘱需同步提供公证书原件扫描件'},
                        {'name': '继承人有效期内身份证', 'detail': '本人身份证正反面清晰彩色扫描件，无遮挡、无涂改'},
                        {'name': '逝者死亡证明', 'detail': '有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡的生效判决书，需加盖出具机构公章'},
                        {'name': '继承人与逝者的亲属关系证明', 'detail': '用于佐证身份关联，开具渠道：户籍地派出所、街道办/村委会、公证处出具的亲属关系公证书'},
                        {'name': '遗嘱效力辅助文件', 'detail': '遗嘱订立全程录像、见证人证言、笔迹鉴定报告等，用于强化遗嘱法律效力'},
                        {'name': '账号归属佐证材料', 'detail': '逝者微信账号注册信息、绑定手机号、实名信息、历史充值记录等，可佐证账号归属'},
                        {'name': '数字遗产明细清单', 'detail': '清晰标注申请处置的资产名称、对应账号、数量、预估价值'},
                        {'name': '接受遗赠声明', 'detail': '由受遗赠人出具，明确表示接受逝者遗赠的数字财产。若受遗赠人为未成年人，可由法定监护人代为签署'}
                    ],
                    'templates': [
                        '数字遗产明细清单',
                        '自书数字遗嘱参考模板',
                        '接受遗赠声明'
                    ]
                },
                {
                    'step': 4,
                    'title': '官方渠道对接——联系客服确认继承流程',
                    'description': '通过微信官方渠道，与客服取得联系，说明您"持有遗嘱且知晓密码"的具体情况，确认最新的材料要求和申请流程。',
                    'contact_channels': True,
                    'materials': [
                        {'name': '微信官方客服专线', 'detail': '95017'},
                        {'name': '线上官方入口', 'detail': '微信APP内「我-设置-帮助与反馈」'},
                        {'name': '线上官方入口', 'detail': '微信/QQ端「腾讯客服」小程序'},
                        {'name': '官方公示的数字遗产处置专属受理邮箱', 'detail': '官方公示的数字遗产处置专属受理邮箱'}
                    ],
                    'communication_items': [
                        {'title': '清晰告知核心信息', 'content': '逝者姓名、身份证号、微信账号、离世时间，继承人姓名、与逝者关系、持有合法有效遗嘱、已知晓账号登录密码、申请合规处置账号内数字遗产'},
                        {'title': '确认关键规则', 'content': '无密码场景的专属受理通道、材料提交的指定方式与格式要求、审核周期、账号归属核验要求、进度查询方式、额外需要补充的个性化材料'}
                    ],
                    'templates': [],
                    'warning': '全程留存沟通凭证：客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号。严禁向客服以外的任何第三方泄露个人与逝者的隐私信息、证明材料。拒绝任何"付费代申请、代破解密码"的第三方服务，避免诈骗。'
                },
                {
                    'step': 5,
                    'title': '正式申请提交与审核跟进',
                    'description': '按官方要求提交全套申请材料，完成正式申请提交，跟进审核进度，处理审核结果。具体操作指引：',
                    'materials': [],
                    'templates': [],
                    'instructions': [
                        '规范提交材料：所有材料按官方要求命名、排序、打包，确保扫描件清晰、信息完整，通过官方确认的专属通道提交，留存提交成功的完整凭证（提交截图、邮件发送记录、受理编号）。',
                        '进度跟进：按官方公示的审核周期定期查询进度，保持预留的手机号、邮箱全程畅通，收到补充材料通知后，在规定时限内按要求补正，留存补正凭证。',
                        '核验配合：按官方指引完成继承人身份二次核验，配合完成账号归属、遗嘱效力的补充确认。',
                        '初审结果处理：审核通过进入资产处置环节，按官方指引完成后续操作；审核不通过第一时间向客服确认驳回的具体原因、补正要求、复核通道，准备补充材料申请复核。'
                    ]
                },
                {
                    'step': 6,
                    'title': '申请被驳回后的兜底法律维权途径',
                    'description': '针对多次复核仍被驳回的申请，通过合法法律途径主张继承权。',
                    'materials': [],
                    'templates': [],
                    'legal_actions': [
                        '非诉维权途径：由律师出具正式律师函，向平台公司主张合法继承权，明确诉求与法律依据，推动诉求解决；向12315平台、工信部、政务服务热线等监管部门提交合规投诉，辅助推进问题解决。',
                        '<div>诉讼维权途径：<br>（1）前置专业法律咨询：委托具备网络虚拟财产、继承纠纷办案经验的专业律师，提供全套申请材料、沟通记录、官方驳回凭证，确认维权方案，完整整理证据链。<br>（2）向有管辖权的人民法院提起继承纠纷诉讼，提交全套证据材料，主张合法继承权，依据法院生效的判决书、调解书，向平台申请执行对应的遗产处置方案。</div>'
                    ]
                }
            ]
        else:
            # 抖音平台专属步骤（根据PDF文件）
            if platform == '抖音':
                steps = [
                    {
                        'step': 1,
                        'title': '明确继承目标',
                        'description': '',
                        'materials_table': True,
                        'materials': [
                            {'name': '账号本身（登录权、使用权等）', 'detail': '根据抖音用户协议及相关政策，账号所有权归北京微播视界科技有限公司所有，用户仅获使用权，且使用权与初始注册人人身身份紧密绑定，平台未允许通过继承转移使用权；逝者未注销账号将自动转为纪念账号，继承人无权获取登录及使用权。', 'extra': '不可继承'},
                            {'name': '账号内虚拟财产', 'detail': '包括抖音钱包余额、抖币、已购买的虚拟商品（如直播打赏道具、会员服务等），此类财产属于用户合法拥有的虚拟财产或债权，符合《民法典》继承编规定，凭合法遗嘱及相关证明按平台流程办理继承。', 'extra': '可依法继承'},
                            {'name': '逝者个人信息（含作品、数据等）', 'detail': '依据《个人信息保护法》及抖音相关政策，逝者近亲属为维护合法权益，可申请复制、下载逝者账号内个人信息（如发布的视频、图文、聊天记录等），超出常规范围的可通过客服沟通处理，平台将提供合法转移路径。', 'extra': '可依法复制/下载/转移'}
                        ],
                        'templates': [],
                        'warning': '【场景合规提示】即使您知晓逝者抖音账号的密码，也严禁利用密码进行任何未经授权或超出继承申请范围的操作，例如：擅自转移大额资金、删除聊天记录等原始数据、冒用身份进行社交或交易。此类行为可能违反平台协议、遗嘱本意及相关法律，可能导致您的继承申请被拒绝，甚至需要承担法律责任。'
                    },
                    {
                        'step': 2,
                        'title': '办理继承权公证',
                        'description': '材料齐全且无误后，前往公证处办理。（详见：《数字遗产继承公证办理流程说明》）',
                        'materials': [
                            {'name': '被继承人死亡证明', 'detail': '医院或派出所等部门出具'},
                            {'name': '继承关系证明', 'detail': '如户口本、结婚证等'},
                            {'name': '继承人身份证明', 'detail': '继承人的身份证、户口簿等有效证件；若部分继承人放弃继承，需提交《放弃继承权声明书》（需公证）'},
                            {'name': '财产凭证', 'detail': '包括抖音钱包余额截图、抖币充值记录、虚拟商品购买订单、抖音支付交易记录等，用于证明账号内虚拟财产的存在与价值'}
                        ],
                        'templates': ['数字遗产继承公证办理流程说明']
                    },
                    {
                        'step': 3,
                        'title': '材料整理与遗嘱效力核验',
                        'description': '提前备齐所有符合平台要求、具备法律效力的申请材料，核验遗嘱法定效力，梳理已知遗产明细，避免因材料不全、不合规被驳回。',
                        'materials': [
                            {'name': '合法有效遗嘱原件', 'detail': '符合《民法典》法定形式，纸质原件彩色扫描件；公证遗嘱需同步提供公证书原件扫描件'},
                            {'name': '继承人有效期内身份证', 'detail': '本人身份证正反面清晰彩色扫描件，无遮挡、无涂改'},
                            {'name': '逝者死亡证明', 'detail': '有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡的生效判决书，需加盖出具机构公章'},
                            {'name': '继承人与逝者的亲属关系证明', 'detail': '用于佐证身份关联，开具渠道：户籍地派出所、街道办/村委会、公证处出具的亲属关系公证书'},
                            {'name': '遗嘱效力辅助文件', 'detail': '遗嘱订立全程录像、见证人证言、笔迹鉴定报告等，用于强化遗嘱法律效力'},
                            {'name': '账号归属佐证材料', 'detail': '逝者抖音账号注册信息、绑定手机号、实名验证信息、历史充值记录、常用登录设备信息等，可佐证账号归属'},
                            {'name': '数字遗产明细清单', 'detail': '清晰标注申请处置的虚拟资产名称（如抖币、钱包余额、虚拟道具等）、对应抖音账号、数量、预估价值；明确申请复制/下载/转移的个人信息范围（如作品、聊天记录、个人资料等）'},
                            {'name': '接受遗赠声明', 'detail': '由受遗赠人出具，明确表示接受逝者遗赠的数字财产。若受遗赠人为未成年人，可由法定监护人代为签署'}
                        ],
                        'templates': [
                            '数字遗产明细清单',
                            '自书数字遗嘱参考模板',
                            '接受遗赠声明'
                        ]
                    },
                    {
                        'step': 4,
                        'title': '官方渠道对接——联系客服确认继承流程',
                        'description': '通过官方渠道对接客服，确认有遗嘱、有密码场景下的有效受理通道、材料要求、申请流程，同步告知继承诉求，留存完整沟通凭证。',
                        'contact_channels': True,
                        'materials': [
                            {'name': '抖音官方客服专线', 'detail': '95152'},
                            {'name': '线上官方入口', 'detail': '抖音APP内「我-≡-我的客服」进入用户反馈界面'},
                            {'name': '线上官方入口', 'detail': '微信/QQ端搜索"抖音"公众号'},
                            {'name': '官方公示的数字遗产处置专属受理邮箱', 'detail': 'feedback@douyin.com'}
                        ],
                        'communication_items': [
                            {'title': '清晰告知核心信息', 'content': '逝者姓名、身份证号、抖音账号、离世时间，继承人姓名、与逝者关系、持有合法有效遗嘱、已知晓账号登录密码、申请合规处置账号内数字遗产'},
                            {'title': '确认关键规则', 'content': '有密码场景的专属受理通道、材料提交的指定方式与格式要求、审核周期、账号归属及遗嘱效力核验要求、个人信息处理的范围限制、虚拟财产提取方式、进度查询方式、额外需要补充的个性化材料'}
                        ],
                        'templates': [],
                        'warning': '全程留存沟通凭证：客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号。严禁向客服以外的任何第三方泄露个人与逝者的隐私信息、证明材料。拒绝任何"付费代申请、代破解密码"的第三方服务，避免诈骗。'
                    },
                    {
                        'step': 5,
                        'title': '正式申请提交与审核跟进',
                        'description': '按官方要求提交全套申请材料，完成正式申请提交，跟进审核进度，处理审核结果。具体操作指引：',
                        'materials': [],
                        'templates': [],
                        'instructions': [
                            '规范提交材料：所有材料按官方要求命名、排序、打包，确保扫描件清晰、信息完整，通过官方确认的专属通道提交，留存提交成功的完整凭证（提交截图、邮件发送记录、受理编号）。',
                            '进度跟进：按官方公示的审核周期定期查询进度，保持预留的手机号、邮箱全程畅通，收到补充材料通知后，在规定时限内按要求补正，留存补正凭证。',
                            '核验配合：按官方指引完成继承人身份二次核验，配合完成账号归属、遗嘱效力的补充确认。',
                            '初审结果处理：审核通过进入资产处置环节，按官方指引完成后续操作；审核不通过第一时间向客服确认驳回的具体原因、补正要求、复核通道，准备补充材料申请复核。'
                        ]
                    },
                    {
                        'step': 6,
                        'title': '申请被驳回后的兜底法律维权途径',
                        'description': '针对多次复核仍被驳回的申请，通过合法法律途径主张继承权。',
                        'materials': [],
                        'templates': [],
                        'legal_actions': [
                            '非诉维权途径：由律师出具正式律师函，向平台公司主张合法继承权，明确诉求与法律依据，推动诉求解决；向12315平台、工信部、政务服务热线等监管部门提交合规投诉，辅助推进问题解决。',
                            '<div>诉讼维权途径：<br>（1）前置专业法律咨询：委托具备网络虚拟财产、继承纠纷办案经验的专业律师，提供全套申请材料、沟通记录、官方驳回凭证，确认维权方案，完整整理证据链。<br>（2）向有管辖权的人民法院提起继承纠纷诉讼，提交全套证据材料，主张合法继承权，依据法院生效的判决书、调解书，向平台申请执行对应的遗产处置方案。</div>'
                        ]
                    }
                ]
            else:
                # QQ平台专属步骤（根据PDF文件）
                if platform == 'QQ':
                    steps = [
                        {
                            'step': 1,
                            'title': '明确继承目标',
                            'description': '',
                            'materials_table': True,
                            'materials': [
                                {'name': '账号本身（登录权、使用权等）', 'detail': '根据《QQ号码使用规则》，QQ账号的所有权归腾讯公司，用户仅获得使用权，且该使用权与初始注册人人身身份紧密绑定，平台未允许通过继承转移使用权。', 'extra': '不可继承'},
                                {'name': '账号内虚拟财产', 'detail': '包括Q币、QQ钱包余额、QQ会员、超级会员、黄钻等已购买的虚拟商品及服务，此类财产属于用户合法拥有的虚拟财产或债权，符合《民法典》继承编规定，凭合法遗嘱及相关证明按平台流程办理继承。', 'extra': '可依法继承'},
                                {'name': '逝者个人信息（含作品、数据等）', 'detail': '依据《个人信息保护法》及腾讯相关政策，逝者近亲属为维护合法权益，可申请复制、下载逝者账号内个人信息（如QQ空间日志、相册、聊天记录等），超出常规范围的可通过客服沟通处理，平台将提供合法转移路径。', 'extra': '可依法复制/下载/转移'}
                            ],
                            'templates': [],
                            'warning': '【场景合规提示】即使您知晓逝者QQ账号的密码，也严禁利用密码进行任何未经授权或超出继承申请范围的操作，例如：擅自转移Q币、删除聊天记录等原始数据、冒用身份进行社交或消费。此类行为可能违反平台协议、遗嘱本意及相关法律，可能导致您的继承申请被拒绝，甚至需要承担法律责任。'
                        },
                        {
                            'step': 2,
                            'title': '办理继承权公证',
                            'description': '材料齐全且无误后，前往公证处办理。（详见：《数字遗产继承公证办理流程说明》）',
                            'materials': [
                                {'name': '被继承人死亡证明', 'detail': '医院或派出所等部门出具'},
                                {'name': '继承关系证明', 'detail': '如户口本、结婚证等'},
                                {'name': '继承人身份证明', 'detail': '继承人的身份证、户口簿等有效证件；若部分继承人放弃继承，需提交《放弃继承权声明书》（需公证）'},
                                {'name': '财产凭证', 'detail': '包括Q币充值记录、QQ钱包余额截图、QQ会员/黄钻等虚拟商品购买订单、QQ支付交易记录等，用于证明账号内虚拟财产的存在与价值'}
                            ],
                            'templates': ['数字遗产继承公证办理流程说明']
                        },
                        {
                            'step': 3,
                            'title': '材料整理与遗嘱效力核验',
                            'description': '提前备齐所有符合平台要求、具备法律效力的申请材料，核验遗嘱法定效力，梳理已知遗产明细，避免因材料不全、不合规被驳回。',
                            'materials': [
                                {'name': '合法有效遗嘱原件', 'detail': '符合《民法典》法定形式，纸质原件彩色扫描件；公证遗嘱需同步提供公证书原件扫描件'},
                                {'name': '继承人有效期内身份证', 'detail': '本人身份证正反面清晰彩色扫描件，无遮挡、无涂改'},
                                {'name': '逝者死亡证明', 'detail': '有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡的生效判决书，需加盖出具机构公章'},
                                {'name': '继承人与逝者的亲属关系证明', 'detail': '用于佐证身份关联，开具渠道：户籍地派出所、街道办/村委会、公证处出具的亲属关系公证书'},
                                {'name': '遗嘱效力辅助文件', 'detail': '遗嘱订立全程录像、见证人证言、笔迹鉴定报告等，用于强化遗嘱法律效力'},
                                {'name': '账号归属佐证材料', 'detail': '逝者QQ账号注册信息、绑定手机号、实名验证信息、历史充值记录、常用登录设备信息等，可佐证账号归属'},
                                {'name': '数字遗产明细清单', 'detail': '清晰标注申请处置的资产名称（如Q币、QQ钱包余额、某游戏道具等）、对应QQ账号、数量、预估价值'},
                                {'name': '接受遗赠声明', 'detail': '由受遗赠人出具，明确表示接受逝者遗赠的数字财产。若受遗赠人为未成年人，可由法定监护人代为签署'}
                            ],
                            'templates': [
                                '数字遗产明细清单',
                                '自书数字遗嘱参考模板',
                                '接受遗赠声明'
                            ]
                        },
                        {
                            'step': 4,
                            'title': '官方渠道对接——联系客服确认继承流程',
                            'description': '通过官方渠道对接客服，确认有遗嘱、有密码场景下的有效受理通道、材料要求、申请流程，同步告知继承诉求，留存完整沟通凭证。',
                            'contact_channels': True,
                            'materials': [
                                {'name': 'QQ官方客服专线', 'detail': '400-881-2345'},
                                {'name': '线上官方入口', 'detail': 'QQAPP内「设置-帮助与反馈」'},
                                {'name': '线上官方入口', 'detail': '腾讯客服官网（https://kf.qq.com/）'},
                                {'name': '线上官方入口', 'detail': '微信/QQ端搜索"腾讯客服"小程序'},
                                {'name': '官方公示的数字遗产处置专属受理邮箱', 'detail': '官方公示的数字遗产处置专属受理邮箱'}
                            ],
                            'communication_items': [
                                {'title': '清晰告知核心信息', 'content': '逝者姓名、身份证号、QQ账号、离世时间，继承人姓名、与逝者关系、持有合法有效遗嘱、已知晓账号登录密码、申请合规处置账号内数字遗产'},
                                {'title': '确认关键规则', 'content': '有密码场景的专属受理通道、材料提交的指定方式与格式要求、审核周期、账号归属及遗嘱效力核验要求、个人信息处理的范围限制、虚拟财产提取方式、进度查询方式、额外需要补充的个性化材料'}
                            ],
                            'templates': [],
                            'warning': '全程留存沟通凭证：客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号。严禁向客服以外的任何第三方泄露个人与逝者的隐私信息、证明材料。拒绝任何"付费代申请、代破解密码"的第三方服务，避免诈骗。'
                        },
                        {
                            'step': 5,
                            'title': '正式申请提交与审核跟进',
                            'description': '按官方要求提交全套申请材料，完成正式申请提交，跟进审核进度，处理审核结果。具体操作指引：',
                            'materials': [],
                            'templates': [],
                            'instructions': [
                                '规范提交材料：所有材料按官方要求命名、排序、打包，确保扫描件清晰、信息完整，通过官方确认的专属通道提交，留存提交成功的完整凭证（提交截图、邮件发送记录、受理编号）。',
                                '进度跟进：按官方公示的审核周期定期查询进度，保持预留的手机号、邮箱全程畅通，收到补充材料通知后，在规定时限内按要求补正，留存补正凭证。',
                                '核验配合：按官方指引完成继承人身份二次核验，配合完成账号归属、遗嘱效力的补充确认。',
                                '初审结果处理：审核通过进入资产处置环节，按官方指引完成后续操作；审核不通过第一时间向客服确认驳回的具体原因、补正要求、复核通道，准备补充材料申请复核。'
                            ]
                        },
                        {
                            'step': 6,
                            'title': '申请被驳回后的兜底法律维权途径',
                            'description': '针对多次复核仍被驳回的申请，通过合法法律途径主张继承权。',
                            'materials': [],
                            'templates': [],
                            'legal_actions': [
                                '非诉维权途径：由律师出具正式律师函，向平台公司主张合法继承权，明确诉求与法律依据，推动诉求解决；向12315平台、工信部、政务服务热线等监管部门提交合规投诉，辅助推进问题解决。',
                                '<div>诉讼维权途径：<br>（1）前置专业法律咨询：委托具备网络虚拟财产、继承纠纷办案经验的专业律师，提供全套申请材料、沟通记录、官方驳回凭证，确认维权方案，完整整理证据链。<br>（2）向有管辖权的人民法院提起继承纠纷诉讼，提交全套证据材料，主张合法继承权，依据法院生效的判决书、调解书，向平台申请执行对应的遗产处置方案。</div>'
                            ]
                        }
                    ]
                else:
                    # 其他平台保持原有逻辑
                    steps = [
                    {
                        'step': 1,
                        'title': '合规准备——材料整理+遗产核验',
                        'description': '备齐全部合规申请材料，在不登录账号的前提下，完成账号内遗产的初步梳理，避免违规操作。',
                        'materials': [
                            '合法有效遗嘱原件（符合《民法典》法定形式，纸质原件彩色扫描件；公证遗嘱需同步提供公证书原件）',
                            '继承人有效期内身份证（本人身份证正反面清晰彩色扫描件，无遮挡、无涂改）',
                            '逝者死亡证明（有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡生效判决书，需加盖出具机构公章）',
                            '继承人与逝者亲属关系证明（开具渠道：户籍地派出所、公证处、街道办/村委会出具的有效亲属关系证明）',
                            '账号归属佐证材料（账号注册信息、绑定手机号、实名信息、历史充值记录，用于佐证账号归属）',
                            '数字遗产明细清单（标注申请处置的资产名称、对应账号、数量、预估价值）'
                        ],
                        'templates': [
                            f'{platform}数字遗产明细清单模板',
                            '自书数字遗嘱参考模板',
                            '继承申请材料整理封面与目录模板'
                        ],
                        'warning': '【场景核心合规提示】⚠️ 即使持有账号登录密码，也严禁擅自登录账号修改信息、转移资产、删除内容，该行为违反《用户服务协议》，可能触发账号风控冻结、回收，同时可能侵犯逝者隐私权，需先通过官方渠道完成合规申请，再按平台指引处置资产。'
                    },
                    {
                        'step': 2,
                        'title': '官方渠道对接——联系客服确认申请规则',
                        'description': f'通过官方渠道对接客服，告知继承诉求，确认当前有效受理通道、材料要求，留存完整沟通凭证。',
                        'materials': [
                            f'{platform}官方客服专线：{info["customer_service"]}',
                            f'线上官方入口：{", ".join(info["official_channels"])}',
                            '客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号'
                        ],
                        'templates': [
                            f'{platform}数字遗产继承客服沟通标准话术模板'
                        ],
                        'communication_points': [
                            f'清晰告知核心信息：逝者身份信息、{platform}账号、离世时间，继承人身份、与逝者关系、持有合法有效遗嘱与账号登录密码、申请合规处置账号内数字遗产',
                            f'确认关键规则：官方受理通道、材料提交方式与格式要求、审核周期、持有密码场景的专属处置流程、进度查询方式'
                        ],
                        'warning': '全程留存沟通凭证：客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号。'
                    },
                    {
                        'step': 3,
                        'title': '正式申请提交与审核跟进',
                        'description': f'按官方要求提交全套申请材料，完成正式申请，跟进审核进度，配合完成核验。',
                        'materials': [
                            f'《{platform}数字遗产继承正式申请表》（按官方要求填写，完整填写逝者与继承人信息、账号信息、遗产明细、申请诉求、合规承诺）',
                            '所有申请材料按官方要求命名、排序、打包',
                            '提交成功凭证（提交截图、邮件记录、受理编号）'
                        ],
                        'templates': [
                            f'{platform}数字遗产继承正式申请表',
                            '继承申请材料补正说明模板'
                        ],
                        'instructions': [
                            '规范提交材料：所有材料按官方要求命名、排序、打包，通过官方指定通道提交，留存提交成功凭证（提交截图、邮件记录、受理编号）',
                            '进度跟进：按官方公示的审核周期定期查询进度，保持预留联系方式畅通，收到补充材料通知后，在时限内完成补正',
                            '核验配合：按官方指引完成继承人身份二次核验，配合完成账号归属、遗嘱效力的补充确认'
                        ]
                    },
                    {
                        'step': 4,
                        'title': '审核结果处置与合规收尾',
                        'description': '根据审核结果完成资产合规处置，或启动兜底维权流程。',
                        'materials': [],
                        'templates': [],
                        'outcome_handling': {
                            'approved': [
                                f'按官方指引与遗嘱分配方案，完成资产合规处置：资金类资产提取划转、虚拟财产合规转移、数字内容合规导出',
                                f'按逝者意愿完成账号后续处置（如账号注销、内容封存），全程留存处置完成凭证',
                                '严禁擅自修改账号密码、转移非遗嘱分配范围内的资产'
                            ],
                            'rejected': [
                                '第一时间确认驳回原因、补正要求、复核通道，补充材料后申请复核',
                                '多次复核仍被驳回的，整理全套申请材料与沟通记录，进入法律维权流程'
                            ]
                        }
                    }
                ]
    elif scenario == 'scenario2':  # 有遗嘱+无密码
        # 微信平台专属步骤
        if platform == '微信':
            steps = [
                {
                    'step': 1,
                    'title': '明确继承目标',
                    'description': '',
                    'materials_table': True,
                    'materials': [
                        {'name': '账号本身（登录权、使用权等）', 'detail': '平台用户协议明确禁止账号继承、转让等', 'extra': '不可继承'},
                        {'name': '账号内虚拟财产', 'detail': '准备基础证明文件，按平台及法律流程办理', 'extra': '可依法继承'}
                    ],
                    'templates': [],
                    'warning': '【场景合规提示】无登录密码的情况下，严禁尝试破解密码、通过短信验证码非授权登录、使用第三方工具登录逝者账号，该行为违反《用户服务协议》与相关法律规定，可能触发账号永久封禁、回收，同时需承担相应的侵权法律责任。'
                },
                {
                    'step': 2,
                    'title': '办理继承权公证',
                    'description': '材料齐全且无误后，前往公证处办理。（详见：《数字遗产继承公证办理流程说明》）',
                    'materials': [
                        {'name': '被继承人死亡证明', 'detail': '医院或派出所等部门出具'},
                        {'name': '继承关系证明', 'detail': '如户口本、结婚证等'},
                        {'name': '继承人身份证明', 'detail': '继承人的身份证、户口簿等有效证件'},
                        {'name': '财产凭证', 'detail': '如微信钱包余额截图、虚拟商品购买记录等，证明账号内虚拟财产的存在与价值'}
                    ],
                    'templates': ['数字遗产继承公证办理流程说明']
                },
                {
                    'step': 3,
                    'title': '材料整理与遗嘱效力核验',
                    'description': '提前备齐所有符合平台要求、具备法律效力的申请材料，核验遗嘱法定效力，梳理已知遗产明细，避免因材料不全、不合规被驳回。',
                    'materials': [
                        {'name': '合法有效遗嘱原件', 'detail': '符合《民法典》法定形式，纸质原件彩色扫描件；公证遗嘱需同步提供公证书原件扫描件'},
                        {'name': '继承人有效期内身份证', 'detail': '本人身份证正反面清晰彩色扫描件，无遮挡、无涂改'},
                        {'name': '逝者死亡证明', 'detail': '有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡的生效判决书，需加盖出具机构公章'},
                        {'name': '继承人与逝者的亲属关系证明', 'detail': '用于佐证身份关联，开具渠道：户籍地派出所、街道办/村委会、公证处出具的亲属关系公证书'},
                        {'name': '遗嘱效力辅助文件', 'detail': '遗嘱订立全程录像、见证人证言、笔迹鉴定报告等，用于强化遗嘱法律效力'},
                        {'name': '账号归属佐证材料', 'detail': '逝者微信账号注册信息、绑定手机号、实名信息、历史充值记录等，可佐证账号归属'},
                        {'name': '数字遗产明细清单', 'detail': '清晰标注申请处置的资产名称、对应账号、数量、预估价值'},
                        {'name': '接受遗赠声明', 'detail': '由受遗赠人出具，明确表示接受逝者遗赠的数字财产。若受遗赠人为未成年人，可由法定监护人代为签署'}
                    ],
                    'templates': [
                        '数字遗产明细清单',
                        '自书数字遗嘱参考模板',
                        '接受遗赠声明'
                    ]
                },
                {
                    'step': 4,
                    'title': '官方渠道对接——联系客服确认继承流程',
                    'description': '通过官方渠道对接客服，确认无密码场景下的有效受理通道、材料要求、申请流程，同步告知继承诉求，留存完整沟通凭证。',
                    'contact_channels': True,
                    'materials': [
                        {'name': '微信官方客服专线', 'detail': '95017'},
                        {'name': '线上官方入口', 'detail': '微信APP内「我-设置-帮助与反馈」'},
                        {'name': '线上官方入口', 'detail': '微信/QQ端「腾讯客服」小程序'},
                        {'name': '官方公示的数字遗产处置专属受理邮箱', 'detail': '官方公示的数字遗产处置专属受理邮箱'}
                    ],
                    'communication_items': [
                        {'title': '清晰告知核心信息', 'content': '逝者姓名、身份证号、微信账号、离世时间，继承人姓名、与逝者关系、持有合法有效遗嘱、无账号登录密码、申请合规处置账号内数字遗产'},
                        {'title': '确认关键规则', 'content': '无密码场景的专属受理通道、材料提交的指定方式与格式要求、审核周期、账号归属核验要求、进度查询方式、额外需要补充的个性化材料'}
                    ],
                    'templates': [],
                    'warning': '全程留存沟通凭证：客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号。严禁向客服以外的任何第三方泄露个人与逝者的隐私信息、证明材料。拒绝任何"付费代申请、代破解密码"的第三方服务，避免诈骗。'
                },
                {
                    'step': 5,
                    'title': '正式申请提交与审核跟进',
                    'description': '按官方要求提交全套申请材料，完成正式申请提交，跟进审核进度，处理审核结果。具体操作指引：',
                    'materials': [],
                    'templates': [],
                    'instructions': [
                        '规范提交材料：所有材料按官方要求命名、排序、打包，确保扫描件清晰、信息完整，通过官方确认的专属通道提交，留存提交成功的完整凭证（提交截图、邮件发送记录、受理编号）。',
                        '进度跟进：按官方公示的审核周期定期查询进度，保持预留的手机号、邮箱全程畅通，收到补充材料通知后，在规定时限内按要求补正，留存补正凭证。',
                        '核验配合：按官方指引完成继承人身份二次核验，配合完成账号归属、遗嘱效力的补充确认。',
                        '初审结果处理：审核通过进入资产处置环节，按官方指引完成后续操作；审核不通过第一时间向客服确认驳回的具体原因、补正要求、复核通道，准备补充材料申请复核。'
                    ]
                },
                {
                    'step': 6,
                    'title': '申请被驳回后的兜底法律维权途径',
                    'description': '针对多次复核仍被驳回的申请，通过合法法律途径主张继承权。',
                    'materials': [],
                    'templates': [],
                    'legal_actions': [
                        '非诉维权途径：由律师出具正式律师函，向平台公司主张合法继承权，明确诉求与法律依据，推动诉求解决；向12315平台、工信部、政务服务热线等监管部门提交合规投诉，辅助推进问题解决。',
                        '<div>诉讼维权途径：<br>（1）前置专业法律咨询：委托具备网络虚拟财产、继承纠纷办案经验的专业律师，提供全套申请材料、沟通记录、官方驳回凭证，确认维权方案，完整整理证据链。<br>（2）向有管辖权的人民法院提起继承纠纷诉讼，提交全套证据材料，主张合法继承权，依据法院生效的判决书、调解书，向平台申请执行对应的遗产处置方案。</div>'
                    ]
                }
            ]
        else:
            # 抖音平台专属步骤（根据PDF文件）
            if platform == '抖音':
                steps = [
                    {
                        'step': 1,
                        'title': '明确继承目标',
                        'description': '',
                        'materials_table': True,
                        'materials': [
                            {'name': '账号本身（登录权、使用权等）', 'detail': '根据抖音用户协议及相关政策，账号所有权归北京微播视界科技有限公司所有，用户仅获使用权，且使用权与初始注册人人身身份紧密绑定，平台未允许通过继承转移使用权；逝者未注销账号将自动转为纪念账号，继承人无权获取登录及使用权。', 'extra': '不可继承'},
                            {'name': '账号内虚拟财产', 'detail': '包括抖音钱包余额、抖币、已购买的虚拟商品（如直播打赏道具、会员服务等），此类财产属于用户合法拥有的虚拟财产或债权，符合《民法典》继承编规定，凭合法遗嘱及相关证明按平台流程办理继承。', 'extra': '可依法继承'},
                            {'name': '逝者个人信息（含作品、数据等）', 'detail': '依据《个人信息保护法》及抖音相关政策，逝者近亲属为维护合法权益，可申请复制、下载逝者账号内个人信息（如发布的视频、图文、聊天记录等），超出常规范围的可通过客服沟通处理，平台将提供合法转移路径。', 'extra': '可依法复制/下载/转移'}
                        ],
                        'templates': [],
                        'warning': '【场景合规提示】无登录密码的情况下，严禁尝试破解密码、通过短信验证码非授权登录、使用第三方工具登录逝者账号。该行为违反《抖音用户服务协议》，可能触发账号永久封禁、回收，同时需承担相应的侵权法律责任；且抖音账号连续超过二个月未登录可能被平台收回，建议尽快启动继承流程。'
                    },
                    {
                        'step': 2,
                        'title': '办理继承权公证',
                        'description': '材料齐全且无误后，前往公证处办理。（详见：《数字遗产继承公证办理流程说明》）',
                        'materials': [
                            {'name': '被继承人死亡证明', 'detail': '医院或派出所等部门出具'},
                            {'name': '继承关系证明', 'detail': '如户口本、结婚证等'},
                            {'name': '继承人身份证明', 'detail': '继承人的身份证、户口簿等有效证件；若部分继承人放弃继承，需提交《放弃继承权声明书》（需公证）'},
                            {'name': '财产凭证', 'detail': '包括抖音钱包余额截图、抖币充值记录、虚拟商品购买订单、抖音支付交易记录等，用于证明账号内虚拟财产的存在与价值'}
                        ],
                        'templates': ['数字遗产继承公证办理流程说明']
                    },
                    {
                        'step': 3,
                        'title': '材料整理与遗嘱效力核验',
                        'description': '提前备齐所有符合平台要求、具备法律效力的申请材料，核验遗嘱法定效力，梳理已知遗产明细，避免因材料不全、不合规被驳回。',
                        'materials': [
                            {'name': '合法有效遗嘱原件', 'detail': '符合《民法典》法定形式，纸质原件彩色扫描件；公证遗嘱需同步提供公证书原件扫描件'},
                            {'name': '继承人有效期内身份证', 'detail': '本人身份证正反面清晰彩色扫描件，无遮挡、无涂改'},
                            {'name': '逝者死亡证明', 'detail': '有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡的生效判决书，需加盖出具机构公章'},
                            {'name': '继承人与逝者的亲属关系证明', 'detail': '用于佐证身份关联，开具渠道：户籍地派出所、街道办/村委会、公证处出具的亲属关系公证书'},
                            {'name': '遗嘱效力辅助文件', 'detail': '遗嘱订立全程录像、见证人证言、笔迹鉴定报告等，用于强化遗嘱法律效力'},
                            {'name': '账号归属佐证材料', 'detail': '逝者抖音账号注册信息、绑定手机号、实名验证信息、历史充值记录、常用登录设备信息等，可佐证账号归属'},
                            {'name': '数字遗产明细清单', 'detail': '清晰标注申请处置的虚拟资产名称（如抖币、钱包余额、虚拟道具等）、对应抖音账号、数量、预估价值；明确申请复制/下载/转移的个人信息范围（如作品、聊天记录、个人资料等）'},
                            {'name': '接受遗赠声明', 'detail': '由受遗赠人出具，明确表示接受逝者遗赠的数字财产。若受遗赠人为未成年人，可由法定监护人代为签署'}
                        ],
                        'templates': [
                            '数字遗产明细清单',
                            '自书数字遗嘱参考模板',
                            '接受遗赠声明'
                        ]
                    },
                    {
                        'step': 4,
                        'title': '官方渠道对接——联系客服确认继承流程',
                        'description': '通过官方渠道对接客服，确认有遗嘱、有密码场景下的有效受理通道、材料要求、申请流程，同步告知继承诉求，留存完整沟通凭证。',
                        'contact_channels': True,
                        'materials': [
                            {'name': '抖音官方客服专线', 'detail': '95152'},
                            {'name': '线上官方入口', 'detail': '抖音APP内「我-≡-我的客服」进入用户反馈界面'},
                            {'name': '线上官方入口', 'detail': '微信/QQ端搜索"抖音"公众号'},
                            {'name': '官方公示的数字遗产处置专属受理邮箱', 'detail': 'feedback@douyin.com'}
                        ],
                        'communication_items': [
                            {'title': '清晰告知核心信息', 'content': '逝者姓名、身份证号、抖音账号、离世时间，继承人姓名、与逝者关系、持有合法有效遗嘱、无账号登录密码、已办理继承权公证、申请处置账号内虚拟遗产及复制/转移个人信息的具体诉求'},
                            {'title': '确认关键规则', 'content': '无密码场景的专属受理通道、材料提交的指定方式与格式要求、审核周期、账号归属及遗嘱效力核验要求、个人信息处理的范围限制、虚拟财产提取方式、进度查询方式、额外需要补充的个性化材料'}
                        ],
                        'templates': [],
                        'warning': '全程留存沟通凭证：客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号。严禁向客服以外的任何第三方泄露个人与逝者的隐私信息、证明材料。拒绝任何"付费代申请、代破解密码"的第三方服务，避免诈骗。'
                    },
                    {
                        'step': 5,
                        'title': '正式申请提交与审核跟进',
                        'description': '按官方要求提交全套申请材料，完成正式申请提交，跟进审核进度，处理审核结果。具体操作指引：',
                        'materials': [],
                        'templates': [],
                        'instructions': [
                            '规范提交材料：所有材料按官方要求命名、排序、打包，确保扫描件清晰、信息完整，通过官方确认的专属通道提交，留存提交成功的完整凭证（提交截图、邮件发送记录、受理编号）。',
                            '进度跟进：按官方公示的审核周期定期查询进度，保持预留的手机号、邮箱全程畅通，收到补充材料通知后，在规定时限内按要求补正，留存补正凭证。',
                            '核验配合：按官方指引完成继承人身份二次核验，配合完成账号归属、遗嘱效力的补充确认。',
                            '初审结果处理：审核通过进入资产处置环节，按官方指引完成后续操作；审核不通过第一时间向客服确认驳回的具体原因、补正要求、复核通道，准备补充材料申请复核。'
                        ]
                    },
                    {
                        'step': 6,
                        'title': '申请被驳回后的兜底法律维权途径',
                        'description': '针对多次复核仍被驳回的申请，通过合法法律途径主张继承权。',
                        'materials': [],
                        'templates': [],
                        'legal_actions': [
                            '非诉维权途径：由律师出具正式律师函，向平台公司主张合法继承权，明确诉求与法律依据，推动诉求解决；向12315平台、工信部、政务服务热线等监管部门提交合规投诉，辅助推进问题解决。',
                            '<div>诉讼维权途径：<br>（1）前置专业法律咨询：委托具备网络虚拟财产、继承纠纷办案经验的专业律师，提供全套申请材料、沟通记录、官方驳回凭证，确认维权方案，完整整理证据链。<br>（2）向有管辖权的人民法院提起继承纠纷诉讼，提交全套证据材料，主张合法继承权，依据法院生效的判决书、调解书，向平台申请执行对应的遗产处置方案。</div>'
                        ]
                    }
                ]
            else:
                # QQ平台专属步骤（根据PDF文件）
                if platform == 'QQ':
                    steps = [
                        {
                            'step': 1,
                            'title': '明确继承目标',
                            'description': '',
                            'materials_table': True,
                            'materials': [
                                {'name': '账号本身（登录权、使用权等）', 'detail': '根据《QQ号码使用规则》，QQ账号的所有权归腾讯公司，用户仅获得使用权，且该使用权与初始注册人人身身份紧密绑定，平台未允许通过继承转移使用权。', 'extra': '不可继承'},
                                {'name': '账号内虚拟财产', 'detail': '包括Q币、QQ钱包余额、QQ会员、超级会员、黄钻等已购买的虚拟商品及服务，此类财产属于用户合法拥有的虚拟财产或债权，符合《民法典》继承编规定，凭合法遗嘱及相关证明按平台流程办理继承。', 'extra': '可依法继承'},
                                {'name': '逝者个人信息（含作品、数据等）', 'detail': '依据《个人信息保护法》及腾讯相关政策，逝者近亲属为维护合法权益，可申请复制、下载逝者账号内个人信息（如QQ空间日志、相册、聊天记录等），超出常规范围的可通过客服沟通处理，平台将提供合法转移路径。', 'extra': '可依法复制/下载/转移'}
                            ],
                            'templates': [],
                            'warning': '【场景合规提示】无登录密码的情况下，严禁尝试破解密码、通过短信验证码非授权登录、使用第三方工具登录逝者账号。该行为违反《QQ号码使用规则》，可能触发账号永久封禁、回收，同时需承担相应的侵权法律责任；且QQ账号长期未登录可能被平台回收，建议尽快启动继承流程。'
                        },
                        {
                            'step': 2,
                            'title': '办理继承权公证',
                            'description': '材料齐全且无误后，前往公证处办理。（详见：《数字遗产继承公证办理流程说明》）',
                            'materials': [
                                {'name': '被继承人死亡证明', 'detail': '医院或派出所等部门出具'},
                                {'name': '继承关系证明', 'detail': '如户口本、结婚证等'},
                                {'name': '继承人身份证明', 'detail': '继承人的身份证、户口簿等有效证件；若部分继承人放弃继承，需提交《放弃继承权声明书》（需公证）'},
                                {'name': '财产凭证', 'detail': '包括Q币充值记录、QQ钱包余额截图、QQ会员/黄钻等虚拟商品购买订单、QQ支付交易记录等，用于证明账号内虚拟财产的存在与价值'}
                            ],
                            'templates': ['数字遗产继承公证办理流程说明']
                        },
                        {
                            'step': 3,
                            'title': '材料整理与遗嘱效力核验',
                            'description': '提前备齐所有符合平台要求、具备法律效力的申请材料，核验遗嘱法定效力，梳理已知遗产明细，避免因材料不全、不合规被驳回。',
                            'materials': [
                                {'name': '合法有效遗嘱原件', 'detail': '符合《民法典》法定形式，纸质原件彩色扫描件；公证遗嘱需同步提供公证书原件扫描件'},
                                {'name': '继承人有效期内身份证', 'detail': '本人身份证正反面清晰彩色扫描件，无遮挡、无涂改'},
                                {'name': '逝者死亡证明', 'detail': '有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡的生效判决书，需加盖出具机构公章'},
                                {'name': '继承人与逝者的亲属关系证明', 'detail': '用于佐证身份关联，开具渠道：户籍地派出所、街道办/村委会、公证处出具的亲属关系公证书'},
                                {'name': '遗嘱效力辅助文件', 'detail': '遗嘱订立全程录像、见证人证言、笔迹鉴定报告等，用于强化遗嘱法律效力'},
                                {'name': '账号归属佐证材料', 'detail': '逝者QQ账号注册信息、绑定手机号、实名验证信息、历史充值记录、常用登录设备信息等，可佐证账号归属'},
                                {'name': '数字遗产明细清单', 'detail': '清晰标注申请处置的资产名称（如Q币、QQ钱包余额、某游戏道具等）、对应QQ账号、数量、预估价值'},
                                {'name': '接受遗赠声明', 'detail': '由受遗赠人出具，明确表示接受逝者遗赠的数字财产。若受遗赠人为未成年人，可由法定监护人代为签署'}
                            ],
                            'templates': [
                                '数字遗产明细清单',
                                '自书数字遗嘱参考模板',
                                '接受遗赠声明'
                            ]
                        },
                        {
                            'step': 4,
                            'title': '官方渠道对接——联系客服确认继承流程',
                            'description': '通过官方渠道对接客服，确认有遗嘱、无密码场景下的有效受理通道、材料要求、申请流程，同步告知继承诉求，留存完整沟通凭证。',
                            'contact_channels': True,
                            'materials': [
                                {'name': 'QQ官方客服专线', 'detail': '400-881-2345'},
                                {'name': '线上官方入口', 'detail': 'QQAPP内「设置-帮助与反馈」'},
                                {'name': '线上官方入口', 'detail': '腾讯客服官网（https://kf.qq.com/）'},
                                {'name': '线上官方入口', 'detail': '微信/QQ端搜索"腾讯客服"小程序'},
                                {'name': '官方公示的数字遗产处置专属受理邮箱', 'detail': '官方公示的数字遗产处置专属受理邮箱'}
                            ],
                            'communication_items': [
                                {'title': '清晰告知核心信息', 'content': '逝者姓名、身份证号、QQ账号、离世时间，继承人姓名、与逝者关系、持有合法有效遗嘱、无账号登录密码、已办理继承权公证、申请处置账号内虚拟遗产及复制/转移个人信息的具体诉求'},
                                {'title': '确认关键规则', 'content': '无密码场景的专属受理通道、材料提交的指定方式与格式要求、审核周期、账号归属及遗嘱效力核验要求、个人信息处理的范围限制、虚拟财产提取方式、进度查询方式、额外需要补充的个性化材料'}
                            ],
                            'templates': [],
                            'warning': '全程留存沟通凭证：客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号。严禁向客服以外的任何第三方泄露个人与逝者的隐私信息、证明材料。拒绝任何"付费代申请、代破解密码"的第三方服务，避免诈骗。'
                        },
                        {
                            'step': 5,
                            'title': '正式申请提交与审核跟进',
                            'description': '按官方要求提交全套申请材料，完成正式申请提交，跟进审核进度，处理审核结果。具体操作指引：',
                            'materials': [],
                            'templates': [],
                            'instructions': [
                                '规范提交材料：所有材料按官方要求命名、排序、打包，确保扫描件清晰、信息完整，通过官方确认的专属通道提交，留存提交成功的完整凭证（提交截图、邮件发送记录、受理编号）。',
                                '进度跟进：按官方公示的审核周期定期查询进度，保持预留的手机号、邮箱全程畅通，收到补充材料通知后，在规定时限内按要求补正，留存补正凭证。',
                                '核验配合：按官方指引完成继承人身份二次核验，配合完成账号归属、遗嘱效力的补充确认。',
                                '初审结果处理：审核通过进入资产处置环节，按官方指引完成后续操作；审核不通过第一时间向客服确认驳回的具体原因、补正要求、复核通道，准备补充材料申请复核。'
                            ]
                        },
                        {
                            'step': 6,
                            'title': '申请被驳回后的兜底法律维权途径',
                            'description': '针对多次复核仍被驳回的申请，通过合法法律途径主张继承权。',
                            'materials': [],
                            'templates': [],
                            'legal_actions': [
                                '非诉维权途径：由律师出具正式律师函，向平台公司主张合法继承权，明确诉求与法律依据，推动诉求解决；向12315平台、工信部、政务服务热线等监管部门提交合规投诉，辅助推进问题解决。',
                               '<div>诉讼维权途径：<br>（1）前置专业法律咨询：委托具备网络虚拟财产、继承纠纷办案经验的专业律师，提供全套申请材料、沟通记录、官方驳回凭证，确认维权方案，完整整理证据链。<br>（2）向有管辖权的人民法院提起继承纠纷诉讼，提交全套证据材料，主张合法继承权，依据法院生效的判决书、调解书，向平台申请执行对应的遗产处置方案。</div>'
                            ]
                        }
                    ]
                else:
                    # 其他平台保持原有逻辑
                    steps = [
                    {
                        'step': 1,
                        'title': '合规准备——材料整理+遗嘱效力核验',
                        'description': '提前备齐所有符合平台要求、具备法律效力的申请材料，核验遗嘱法定效力，梳理已知遗产明细，避免因材料不全、不合规被驳回。',
                        'materials': [
                            '合法有效遗嘱原件（符合《民法典》法定形式，纸质原件彩色扫描件；公证遗嘱需同步提供公证书原件扫描件）',
                            '继承人有效期内身份证（本人身份证正反面清晰彩色扫描件，无遮挡、无涂改）',
                            '逝者死亡证明（有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡的生效判决书，需加盖出具机构公章）',
                            '继承人与逝者的亲属关系证明（用于佐证身份关联，开具渠道：户籍地派出所、街道办/村委会、公证处出具的亲属关系公证书）',
                            '遗嘱效力辅助文件（遗嘱订立全程录像、见证人证言、笔迹鉴定报告等，用于强化遗嘱法律效力）',
                            f'账号归属佐证材料（逝者{platform}账号注册信息、绑定手机号、实名信息、历史充值记录等，可佐证账号归属）',
                            '数字遗产明细清单（清晰标注申请处置的资产名称、对应账号、数量、预估价值）'
                        ],
                        'templates': [
                            f'{platform}数字遗产明细清单模板',
                            '自书数字遗嘱参考模板',
                            '继承申请材料整理封面与目录模板'
                        ],
                        'warning': '【场景核心合规提示】⚠️ 无登录密码的情况下，严禁尝试破解密码、通过短信验证码非授权登录、使用第三方工具登录逝者账号，该行为违反《用户服务协议》与相关法律规定，可能触发账号永久封禁、回收，同时需承担相应的侵权法律责任。'
                    },
                    {
                        'step': 2,
                        'title': '官方渠道对接——联系客服确认继承流程',
                        'description': f'通过官方渠道对接客服，确认无密码场景下的有效受理通道、材料要求、申请流程，同步告知继承诉求，留存完整沟通凭证。',
                        'materials': [
                            f'{platform}官方客服专线：{info["customer_service"]}',
                            f'线上官方入口：{", ".join(info["official_channels"])}',
                            '官方公示的数字遗产处置专属受理邮箱'
                        ],
                        'templates': [
                            f'{platform}数字遗产继承客服沟通标准话术模板'
                        ],
                        'communication_points': [
                            f'清晰告知核心信息：逝者姓名、身份证号、{platform}账号、离世时间，继承人姓名、与逝者关系、持有合法有效遗嘱、无账号登录密码、申请合规处置账号内数字遗产',
                            f'确认关键规则：无密码场景的专属受理通道、材料提交的指定方式与格式要求、审核周期、账号归属核验要求、进度查询方式、额外需要补充的个性化材料'
                        ],
                        'warning': '全程留存沟通凭证：客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号。严禁向客服以外的任何第三方泄露个人与逝者的隐私信息、证明材料。拒绝任何"付费代申请、代破解密码"的第三方服务，避免诈骗。'
                    },
                    {
                        'step': 3,
                        'title': '正式申请提交与审核跟进',
                        'description': f'按官方要求提交全套申请材料，完成正式申请提交，跟进审核进度，处理审核结果。',
                        'materials': [
                            f'《{platform}数字遗产继承正式申请表》（核心内容包括逝者与继承人的身份信息、{platform}账号信息、申请处置的遗产明细、申请诉求、继承人真实性承诺与法律责任声明）',
                            '所有材料按官方要求命名、排序、打包（确保扫描件清晰、信息完整）',
                            '提交成功的完整凭证（提交截图、邮件发送记录、受理编号）'
                        ],
                        'templates': [
                            f'{platform}数字遗产继承正式申请表',
                            '继承申请材料补正说明模板'
                        ],
                        'instructions': [
                            '规范提交材料：所有材料按官方要求命名、排序、打包，确保扫描件清晰、信息完整，通过官方确认的专属通道提交，留存提交成功的完整凭证（提交截图、邮件发送记录、受理编号）',
                            '进度跟进：按官方公示的审核周期定期查询进度，保持预留的手机号、邮箱全程畅通，收到补充材料通知后，在规定时限内按要求补正，留存补正凭证',
                            '核验配合：按官方指引完成继承人身份二次核验，配合完成账号归属、遗嘱效力的补充确认',
                            '初审结果处理：审核通过进入资产处置环节，按官方指引完成后续操作；审核不通过第一时间向客服确认驳回的具体原因、补正要求、复核通道，准备补充材料申请复核'
                        ]
                    },
                    {
                        'step': 4,
                        'title': '申请被驳回后的兜底法律维权途径',
                        'description': '针对多次复核仍被驳回的申请，通过合法法律途径主张继承权。',
                        'materials': [],
                        'templates': [],
                        'legal_actions': [
                            '前置专业法律咨询：委托具备网络虚拟财产、继承纠纷办案经验的专业律师，提供全套申请材料、沟通记录、官方驳回凭证，确认维权方案，完整整理证据链',
                            '非诉维权途径：由律师出具正式律师函，向平台公司主张合法继承权，明确诉求与法律依据，推动诉求解决；向12315平台、工信部、政务服务热线等监管部门提交合规投诉，辅助推进问题解决',
                            '诉讼维权途径：向有管辖权的人民法院提起继承纠纷诉讼，提交全套证据材料，主张合法继承权，依据法院生效的判决书、调解书，向平台申请执行对应的遗产处置方案'
                        ]
                    }
                ]
    elif scenario == 'scenario3':  # 无遗嘱+无密码（法定继承）
        # 微信平台专属步骤（根据PDF文件）
        if platform == '微信':
            steps = [
                {
                    'step': 1,
                    'title': '明确继承目标',
                    'description': '',
                    'materials_table': True,
                    'materials': [
                        {'name': '账号本身（登录权、使用权等）', 'detail': '平台用户协议明确禁止账号继承、转让等', 'extra': '不可继承'},
                        {'name': '账号内虚拟财产', 'detail': '准备基础证明文件，按平台及法律流程办理', 'extra': '可依法继承'}
                    ],
                    'templates': [],
                    'warning': '【场景合规提示】在无法知晓密码的情况下，绝对禁止尝试通过猜测密码、短信验证、第三方软件等任何非授权方式登录逝者微信账号。此类行为不仅违反微信用户协议，导致账号被永久封禁，还可能涉及侵犯公民个人信息等违法行为，需承担相应法律责任。一切操作必须通过官方合规渠道进行。'
                },
                {
                    'step': 2,
                    'title': '办理继承权公证',
                    'description': '材料齐全且无误后，前往公证处办理。（详见：《数字遗产继承公证办理流程说明》）',
                    'materials': [
                        {'name': '被继承人死亡证明', 'detail': '医院或派出所等部门出具'},
                        {'name': '继承关系证明', 'detail': '如户口本、结婚证等'},
                        {'name': '继承人身份证明', 'detail': '继承人的身份证、户口簿等有效证件'},
                        {'name': '财产凭证', 'detail': '如微信钱包余额截图、虚拟商品购买记录等，证明账号内虚拟财产的存在与价值'},
                        {'name': '全体继承人关于遗产分割的协议', 'detail': '所有法定继承人需共同到场，就如何分配逝者的数字遗产达成书面一致意见，并在公证员面前签字确认'}
                    ],
                    'templates': ['数字遗产继承公证办理流程说明']
                },
                {
                    'step': 3,
                    'title': '材料整理与遗嘱效力核验',
                    'description': '提前备齐所有符合平台要求、具备法律效力的申请材料，核验遗嘱法定效力，梳理已知遗产明细，避免因材料不全、不合规被驳回。',
                    'materials': [
                        {'name': '法定继承公证书', 'detail': '公证处出具的《继承权公证书》原件及清晰彩色扫描件'},
                        {'name': '全体继承人有效期内身份证', 'detail': '本人身份证正反面清晰彩色扫描件，无遮挡、无涂改'},
                        {'name': '逝者死亡证明', 'detail': '有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡的生效判决书，需加盖出具机构公章'},
                        {'name': '继承人与逝者的亲属关系证明', 'detail': '用于佐证身份关联，开具渠道：户籍地派出所、街道办/村委会、公证处出具的亲属关系公证书'},
                        {'name': '账号归属佐证材料', 'detail': '逝者微信账号注册信息、绑定手机号、实名信息、历史充值记录等，可佐证账号归属'},
                        {'name': '全体继承人签署的申请委托书', 'detail': '由一位或几位继承人作为代表办理申请，需要其他所有继承人签署委托书，明确授权代表人办理相关事宜'},
                        {'name': '数字遗产明细清单', 'detail': '清晰标注申请处置的资产名称、对应账号、数量、预估价值'},
                        {'name': '接受遗赠声明', 'detail': '由受遗赠人出具，明确表示接受逝者遗赠的数字财产。若受遗赠人为未成年人，可由法定监护人代为签署'}
                    ],
                    'templates': [
                        '数字遗产明细清单',
                        '自书数字遗嘱参考模板',
                        '接受遗赠声明'
                    ]
                },
                {
                    'step': 4,
                    'title': '官方渠道对接——联系客服确认继承流程',
                    'description': '通过官方渠道对接客服，确认无密码场景下的有效受理通道、材料要求、申请流程，同步告知继承诉求，留存完整沟通凭证。',
                    'contact_channels': True,
                    'materials': [
                        {'name': '微信官方客服专线', 'detail': '95017'},
                        {'name': '线上官方入口', 'detail': '微信APP内「我-设置-帮助与反馈」'},
                        {'name': '线上官方入口', 'detail': '微信/QQ端「腾讯客服」小程序'},
                        {'name': '官方公示的数字遗产处置专属受理邮箱', 'detail': '官方公示的数字遗产处置专属受理邮箱'}
                    ],
                    'communication_items': [
                        {'title': '沟通事项', 'content': '"您好，我需要申请继承已故XXX的微信账户内的资产。情况是：我们没有遗嘱，也不知道账号密码。我们已经办理好了所有法定继承人的《继承权公证书》。请问针对这种情况，我应该向哪个专用邮箱或通道提交材料？具体的材料格式和审核流程是怎样的？"'},
                        {'title': '沟通纪律', 'content': '（1）全程记录客服工号、保存好通话录音或在线聊天记录。（2）仅通过官方确认的渠道提交材料，切勿相信任何"付费加速"、"内部代办"等第三方服务，谨防诈骗。（3）妥善保管所有个人和逝者的隐私材料，不泄露给任何非官方人员。'},
                        {'title': '确认关键规则', 'content': '无密码场景的专属受理通道、材料提交的指定方式与格式要求、审核周期、账号归属核验要求、进度查询方式、额外需要补充的个性化材料'}
                    ],
                    'templates': [],
                    'warning': '全程留存沟通凭证：客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号。严禁向客服以外的任何第三方泄露个人与逝者的隐私信息、证明材料。拒绝任何"付费代申请、代破解密码"的第三方服务，避免诈骗。'
                },
                {
                    'step': 5,
                    'title': '正式申请提交与审核跟进',
                    'description': '按官方要求提交全套申请材料，完成正式申请提交，跟进审核进度，处理审核结果。具体操作指引：',
                    'materials': [],
                    'templates': [],
                    'instructions': [
                        '规范提交材料：所有材料按官方要求命名、排序、打包，确保扫描件清晰、信息完整，通过官方确认的专属通道提交，留存提交成功的完整凭证（提交截图、邮件发送记录、受理编号）。',
                        '进度跟进：按官方公示的审核周期定期查询进度，保持预留的手机号、邮箱全程畅通，收到补充材料通知后，在规定时限内按要求补正，留存补正凭证。',
                        '核验配合：按官方指引完成继承人身份二次核验，配合完成账号归属、遗嘱效力的补充确认。',
                        '初审结果处理：审核通过进入资产处置环节，按官方指引完成后续操作；审核不通过第一时间向客服确认驳回的具体原因、补正要求、复核通道，准备补充材料申请复核。'
                    ]
                },
                {
                    'step': 6,
                    'title': '申请被驳回后的兜底法律维权途径',
                    'description': '针对多次复核仍被驳回的申请，通过合法法律途径主张继承权。',
                    'materials': [],
                    'templates': [],
                    'legal_actions': [
                        '非诉维权途径：由律师出具正式律师函，向平台公司主张合法继承权，明确诉求与法律依据，推动诉求解决；向12315平台、工信部、政务服务热线等监管部门提交合规投诉，辅助推进问题解决。',
                        '<div>诉讼维权途径：<br>（1）前置专业法律咨询：委托具备网络虚拟财产、继承纠纷办案经验的专业律师，提供全套申请材料、沟通记录、官方驳回凭证，确认维权方案，完整整理证据链。<br>（2）向有管辖权的人民法院提起继承纠纷诉讼，提交全套证据材料，主张合法继承权，依据法院生效的判决书、调解书，向平台申请执行对应的遗产处置方案。</div>'
                    ]
                }
            ]
        else:
            # 抖音平台专属步骤（根据PDF文件）
            if platform == '抖音':
                steps = [
                    {
                        'step': 1,
                        'title': '明确继承目标',
                        'description': '',
                        'materials_table': True,
                        'materials': [
                            {'name': '账号本身（登录权、使用权等）', 'detail': '根据抖音用户协议及相关政策，账号所有权归北京微播视界科技有限公司所有，用户仅获使用权，且使用权与初始注册人人身身份紧密绑定，平台未允许通过继承转移使用权；逝者未注销账号将自动转为纪念账号，继承人无权获取登录及使用权。', 'extra': '不可继承'},
                            {'name': '账号内虚拟财产', 'detail': '包括抖音钱包余额、抖币、已购买的虚拟商品（如直播打赏道具、会员服务等），此类财产属于用户合法拥有的虚拟财产或债权，符合《民法典》继承编规定，按法定继承规则办理继承。', 'extra': '可依法继承'},
                            {'name': '逝者个人信息（含作品、数据等）', 'detail': '依据《个人信息保护法》及抖音相关政策，逝者近亲属为维护合法权益，可申请复制、下载逝者账号内个人信息（如发布的视频、图文、聊天记录等），超出常规范围的可通过客服沟通处理，平台将提供合法转移路径。', 'extra': '可依法复制/下载/转移'}
                        ],
                        'templates': [],
                        'warning': '【场景合规提示】无遗嘱、无登录密码的情况下，严禁尝试破解密码、通过短信验证码非授权登录、使用第三方工具登录逝者账号。该行为违反《抖音用户服务协议》，可能触发账号永久封禁、回收，同时需承担相应的侵权法律责任；且抖音账号连续超过二个月未登录可能被平台收回，建议尽快启动继承流程。'
                    },
                    {
                        'step': 2,
                        'title': '办理继承权公证',
                        'description': '材料齐全且无误后，前往公证处办理。（详见：《数字遗产继承公证办理流程说明》）',
                        'materials': [
                            {'name': '被继承人死亡证明', 'detail': '医院或派出所等部门出具'},
                            {'name': '继承关系证明', 'detail': '如户口本、结婚证等'},
                            {'name': '继承人身份证明', 'detail': '继承人的身份证、户口簿等有效证件；若部分继承人放弃继承，需提交《放弃继承权声明书》（需公证）'},
                            {'name': '财产凭证', 'detail': '包括抖音钱包余额截图、抖币充值记录、虚拟商品购买订单、抖音支付交易记录等，用于证明账号内虚拟财产的存在与价值'}
                        ],
                        'templates': ['数字遗产继承公证办理流程说明']
                    },
                    {
                        'step': 3,
                        'title': '材料整理与遗嘱效力核验',
                        'description': '提前备齐所有符合平台要求、具备法律效力的申请材料，核验遗嘱法定效力，梳理已知遗产明细，避免因材料不全、不合规被驳回。',
                        'materials': [
                            {'name': '法定继承公证书', 'detail': '公证处出具的《继承权公证书》原件及清晰彩色扫描件'},
                            {'name': '全体继承人有效期内身份证', 'detail': '本人身份证正反面清晰彩色扫描件，无遮挡、无涂改'},
                            {'name': '逝者死亡证明', 'detail': '有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡的生效判决书，需加盖出具机构公章'},
                            {'name': '继承人与逝者的亲属关系证明', 'detail': '用于佐证身份关联，开具渠道：户籍地派出所、街道办/村委会、公证处出具的亲属关系公证书'},
                            {'name': '账号归属佐证材料', 'detail': '逝者抖音账号注册信息、绑定手机号、实名验证信息、历史充值记录、常用登录设备信息等，可佐证账号归属'},
                            {'name': '全体继承人签署的申请委托书', 'detail': '由一位或几位继承人作为代表办理申请，需要其他所有继承人签署委托书，明确授权代表人办理相关事宜'},
                            {'name': '数字遗产明细清单', 'detail': '清晰标注申请处置的虚拟资产名称（如抖币、钱包余额、虚拟道具等）、对应抖音账号、数量、预估价值；明确申请复制/下载/转移的个人信息范围（如作品、聊天记录、个人资料等）'},
                            {'name': '接受遗赠声明', 'detail': '由受遗赠人出具，明确表示接受逝者遗赠的数字财产。若受遗赠人为未成年人，可由法定监护人代为签署'}
                        ],
                        'templates': [
                            '数字遗产明细清单',
                            '自书数字遗嘱参考模板',
                            '接受遗赠声明'
                        ]
                    },
                    {
                        'step': 4,
                        'title': '官方渠道对接——联系客服确认继承流程',
                        'description': '通过官方渠道对接客服，确认无遗嘱、无密码场景下的有效受理通道、材料要求、申请流程，同步告知继承诉求，留存完整沟通凭证。',
                        'contact_channels': True,
                        'materials': [
                            {'name': '抖音官方客服专线', 'detail': '95152'},
                            {'name': '线上官方入口', 'detail': '抖音APP内「我-≡-我的客服」进入用户反馈界面'},
                            {'name': '线上官方入口', 'detail': '微信/QQ端搜索"抖音"公众号'},
                            {'name': '官方公示的数字遗产处置专属受理邮箱', 'detail': 'feedback@douyin.com'}
                        ],
                        'communication_items': [
                            {'title': '清晰告知核心信息', 'content': '逝者姓名、身份证号、抖音账号、离世时间，继承人姓名、与逝者关系、无有效遗嘱、无账号登录密码、已办理继承权公证、申请按法定继承规则处置账号内虚拟遗产及复制/转移个人信息的具体诉求'},
                            {'title': '确认关键规则', 'content': '无遗嘱、无密码场景的专属受理通道、材料提交的指定方式与格式要求、审核周期、账号归属及继承权核验要求、个人信息处理的范围限制、虚拟财产提取方式、进度查询方式、额外需要补充的个性化材料'}
                        ],
                        'templates': [],
                        'warning': '全程留存沟通凭证：客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号。严禁向客服以外的任何第三方泄露个人与逝者的隐私信息、证明材料。拒绝任何"付费代申请、代破解密码"的第三方服务，避免诈骗。'
                    },
                    {
                        'step': 5,
                        'title': '正式申请提交与审核跟进',
                        'description': '按官方要求提交全套申请材料，完成正式申请提交，跟进审核进度，处理审核结果。具体操作指引：',
                        'materials': [],
                        'templates': [],
                        'instructions': [
                            '规范提交材料：所有材料按官方要求命名、排序、打包，确保扫描件清晰、信息完整，通过官方确认的专属通道提交，留存提交成功的完整凭证（提交截图、邮件发送记录、受理编号）。',
                            '进度跟进：按官方公示的审核周期定期查询进度，保持预留的手机号、邮箱全程畅通，收到补充材料通知后，在规定时限内按要求补正，留存补正凭证。',
                            '核验配合：按官方指引完成继承人身份二次核验，配合完成账号归属、继承权的补充确认。',
                            '初审结果处理：审核通过进入资产处置环节，按官方指引完成后续操作；审核不通过第一时间向客服确认驳回的具体原因、补正要求、复核通道，准备补充材料申请复核。'
                        ]
                    },
                    {
                        'step': 6,
                        'title': '申请被驳回后的兜底法律维权途径',
                        'description': '针对多次复核仍被驳回的申请，通过合法法律途径主张继承权。',
                        'materials': [],
                        'templates': [],
                        'legal_actions': [
                            '非诉维权途径：由律师出具正式律师函，向平台公司主张合法继承权，明确诉求与法律依据，推动诉求解决；向12315平台、工信部、政务服务热线等监管部门提交合规投诉，辅助推进问题解决。',
                           '<div>诉讼维权途径：<br>（1）前置专业法律咨询：委托具备网络虚拟财产、继承纠纷办案经验的专业律师，提供全套申请材料、沟通记录、官方驳回凭证，确认维权方案，完整整理证据链。<br>（2）向有管辖权的人民法院提起继承纠纷诉讼，提交全套证据材料，主张合法继承权，依据法院生效的判决书、调解书，向平台申请执行对应的遗产处置方案。</div>'
                        ]
                    }
                ]
            else:
                # QQ平台专属步骤（根据PDF文件）
                if platform == 'QQ':
                    steps = [
                        {
                            'step': 1,
                            'title': '明确继承目标',
                            'description': '',
                            'materials_table': True,
                            'materials': [
                                {'name': '账号本身（登录权、使用权等）', 'detail': '根据《QQ号码使用规则》，QQ账号的所有权归腾讯公司，用户仅获得使用权，且该使用权与初始注册人人身身份紧密绑定，平台未允许通过继承转移使用权。', 'extra': '不可继承'},
                                {'name': '账号内虚拟财产', 'detail': '包括Q币、QQ钱包余额、QQ会员、超级会员、黄钻等已购买的虚拟商品及服务，此类财产属于用户合法拥有的虚拟财产或债权，符合《民法典》继承编规定，按法定继承规则办理继承。', 'extra': '可依法继承'},
                                {'name': '逝者个人信息（含作品、数据等）', 'detail': '依据《个人信息保护法》及腾讯相关政策，逝者近亲属为维护合法权益，可申请复制、下载逝者账号内个人信息（如QQ空间日志、相册、聊天记录等），超出常规范围的可通过客服沟通处理，平台将提供合法转移路径。', 'extra': '可依法复制/下载/转移'}
                            ],
                            'templates': [],
                            'warning': '【场景合规提示】无遗嘱、无登录密码的情况下，严禁尝试破解密码、通过短信验证码非授权登录、使用第三方工具登录逝者账号。该行为违反《QQ号码使用规则》，可能触发账号永久封禁、回收，同时需承担相应的侵权法律责任；且QQ账号长期未登录可能被平台回收，建议尽快启动继承流程。'
                        },
                        {
                            'step': 2,
                            'title': '办理继承权公证',
                            'description': '材料齐全且无误后，前往公证处办理。（详见：《数字遗产继承公证办理流程说明》）',
                            'materials': [
                                {'name': '被继承人死亡证明', 'detail': '医院或派出所等部门出具'},
                                {'name': '继承关系证明', 'detail': '如户口本、结婚证等'},
                                {'name': '继承人身份证明', 'detail': '继承人的身份证、户口簿等有效证件；若部分继承人放弃继承，需提交《放弃继承权声明书》（需公证）'},
                                {'name': '财产凭证', 'detail': '包括Q币充值记录、QQ钱包余额截图、QQ会员/黄钻等虚拟商品购买订单、QQ支付交易记录等，用于证明账号内虚拟财产的存在与价值'}
                            ],
                            'templates': ['数字遗产继承公证办理流程说明']
                        },
                        {
                            'step': 3,
                            'title': '材料整理与遗嘱效力核验',
                            'description': '提前备齐所有符合平台要求、具备法律效力的申请材料，核验遗嘱法定效力，梳理已知遗产明细，避免因材料不全、不合规被驳回。',
                            'materials': [
                                {'name': '法定继承公证书', 'detail': '公证处出具的《继承权公证书》原件及清晰彩色扫描件'},
                                {'name': '全体继承人有效期内身份证', 'detail': '本人身份证正反面清晰彩色扫描件，无遮挡、无涂改'},
                                {'name': '逝者死亡证明', 'detail': '有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡的生效判决书，需加盖出具机构公章'},
                                {'name': '继承人与逝者的亲属关系证明', 'detail': '用于佐证身份关联，开具渠道：户籍地派出所、街道办/村委会、公证处出具的亲属关系公证书'},
                                {'name': '账号归属佐证材料', 'detail': '逝者QQ账号注册信息、绑定手机号、实名验证信息、历史充值记录、常用登录设备信息等，可佐证账号归属'},
                                {'name': '全体继承人签署的申请委托书', 'detail': '由一位或几位继承人作为代表办理申请，需要其他所有继承人签署委托书，明确授权代表人办理相关事宜'},
                                {'name': '数字遗产明细清单', 'detail': '清晰标注申请处置的资产名称（如Q币、QQ钱包余额、某游戏道具等）、对应QQ账号、数量、预估价值'},
                                {'name': '接受遗赠声明', 'detail': '由受遗赠人出具，明确表示接受逝者遗赠的数字财产。若受遗赠人为未成年人，可由法定监护人代为签署'}
                            ],
                            'templates': [
                                '数字遗产明细清单',
                                '自书数字遗嘱参考模板',
                                '接受遗赠声明'
                            ]
                        },
                        {
                            'step': 4,
                            'title': '官方渠道对接——联系客服确认继承流程',
                            'description': '通过官方渠道对接客服，确认无遗嘱、无密码场景下的有效受理通道、材料要求、申请流程，同步告知继承诉求，留存完整沟通凭证。',
                            'contact_channels': True,
                            'materials': [
                                {'name': 'QQ官方客服专线', 'detail': '400-881-2345'},
                                {'name': '线上官方入口', 'detail': 'QQAPP内「设置-帮助与反馈」'},
                                {'name': '线上官方入口', 'detail': '腾讯客服官网（https://kf.qq.com/）'},
                                {'name': '线上官方入口', 'detail': '微信/QQ端搜索"腾讯客服"小程序'},
                                {'name': '官方公示的数字遗产处置专属受理邮箱', 'detail': '官方公示的数字遗产处置专属受理邮箱'}
                            ],
                            'communication_items': [
                                {'title': '清晰告知核心信息', 'content': '逝者姓名、身份证号、QQ账号、离世时间，继承人姓名、与逝者关系、无有效遗嘱、无账号登录密码、已办理继承权公证、申请按法定继承规则处置账号内虚拟遗产及复制/转移个人信息的具体诉求'},
                                {'title': '确认关键规则', 'content': '无遗嘱、无密码场景的专属受理通道、材料提交的指定方式与格式要求、审核周期、账号归属及继承权核验要求、个人信息处理的范围限制、虚拟财产提取方式、进度查询方式、额外需要补充的个性化材料'}
                            ],
                            'templates': [],
                            'warning': '全程留存沟通凭证：客服工号、通话录音、线上聊天记录、邮件往来记录、受理回执编号。严禁向客服以外的任何第三方泄露个人与逝者的隐私信息、证明材料。拒绝任何"付费代申请、代破解密码"的第三方服务，避免诈骗。'
                        },
                        {
                            'step': 5,
                            'title': '正式申请提交与审核跟进',
                            'description': '按官方要求提交全套申请材料，完成正式申请提交，跟进审核进度，处理审核结果。具体操作指引：',
                            'materials': [],
                            'templates': [],
                            'instructions': [
                                '规范提交材料：所有材料按官方要求命名、排序、打包，确保扫描件清晰、信息完整，通过官方确认的专属通道提交，留存提交成功的完整凭证（提交截图、邮件发送记录、受理编号）。',
                                '进度跟进：按官方公示的审核周期定期查询进度，保持预留的手机号、邮箱全程畅通，收到补充材料通知后，在规定时限内按要求补正，留存补正凭证。',
                                '核验配合：按官方指引完成继承人身份二次核验，配合完成账号归属、继承权的补充确认。',
                                '初审结果处理：审核通过进入资产处置环节，按官方指引完成后续操作；审核不通过第一时间向客服确认驳回的具体原因、补正要求、复核通道，准备补充材料申请复核。'
                            ]
                        },
                        {
                            'step': 6,
                            'title': '申请被驳回后的兜底法律维权途径',
                            'description': '针对多次复核仍被驳回的申请，通过合法法律途径主张继承权。',
                            'materials': [],
                            'templates': [],
                            'legal_actions': [
                                '非诉维权途径：由律师出具正式律师函，向平台公司主张合法继承权，明确诉求与法律依据，推动诉求解决；向12315平台、工信部、政务服务热线等监管部门提交合规投诉，辅助推进问题解决。',
                               '<div>诉讼维权途径：<br>（1）前置专业法律咨询：委托具备网络虚拟财产、继承纠纷办案经验的专业律师，提供全套申请材料、沟通记录、官方驳回凭证，确认维权方案，完整整理证据链。<br>（2）向有管辖权的人民法院提起继承纠纷诉讼，提交全套证据材料，主张合法继承权，依据法院生效的判决书、调解书，向平台申请执行对应的遗产处置方案。</div>'
                            ]
                        }
                    ]
                else:
                    # 其他平台保持原有逻辑
                    steps = [
                    {
                        'step': 1,
                        'title': '合规准备——继承资格确认+全套材料整理',
                        'description': '确认法定继承资格，备齐全部符合平台要求的法定继承申请材料，避免因资格不符、材料不全被驳回。',
                        'materials': [
                            '法定继承人有效期内身份证（所有申请继承的权利人身份证正反面清晰彩色扫描件）',
                            '逝者死亡证明（有效类型：医院出具的《死亡医学证明》、户籍地派出所出具的户籍注销证明、法院宣告死亡生效判决书，需加盖出具机构公章）',
                            '法定继承亲属关系证明（可证明继承人与逝者法定继承关系的文件，开具渠道：户籍地派出所、公证处出具的继承权公证书、街道办/村委会出具的亲属关系证明）',
                            '继承资格与分配方案文件（多名继承人的，需提供全体继承人签字确认的继承分配方案、授权委托书（明确牵头申请人与委托权限））',
                            f'账号归属佐证材料（逝者{platform}账号注册信息、绑定手机号、实名信息、历史充值记录，用于佐证账号归属）',
                            '数字遗产明细清单（已知的账号内资产明细，标注资产名称、预估价值）'
                        ],
                        'templates': [
                            '法定继承亲属关系证明办理指引模板',
                            '多继承人继承分配方案与授权委托书模板',
                            f'{platform}数字遗产明细清单模板'
                        ],
                        'warning': '【场景核心合规提示】\n1. 法定继承资格严格按照《民法典》规定：继承开始后，由第一顺序继承人（配偶、子女、父母）继承，无第一顺序继承人的，由第二顺序继承人（兄弟姐妹、祖父母、外祖父母）继承\n2. 多名法定继承人共同申请的，需提供全体继承人的身份证明、授权委托书，明确牵头申请人与分配方案，避免因继承资格问题驳回申请'
                    },
                    {
                        'step': 2,
                        'title': '官方渠道对接——联系客服确认法定继承流程',
                        'description': f'通过官方渠道对接客服，告知法定继承诉求，确认平台法定继承的专属受理规则、材料要求、申请通道。',
                        'materials': [
                            f'{platform}官方客服专线：{info["customer_service"]}',
                            f'线上官方入口：{", ".join(info["official_channels"])}',
                            '客服工号、通话录音、线上聊天记录、受理回执编号'
                        ],
                        'templates': [
                            f'{platform}平台法定继承场景客服沟通标准话术模板'
                        ],
                        'communication_points': [
                            f'清晰告知核心信息：逝者身份信息、{platform}账号、离世时间，法定继承人身份、与逝者的继承关系、无有效遗嘱、申请按法定继承规则处置账号内数字遗产',
                            f'确认关键规则：平台法定继承的受理要求、材料提交方式、审核周期、多继承人场景的专属要求、进度查询方式'
                        ],
                        'warning': '全程留存沟通凭证，多名继承人的需明确统一的对接人与联系方式，避免信息混乱。'
                    },
                    {
                        'step': 3,
                        'title': '正式申请提交与审核跟进',
                        'description': f'按官方要求提交法定继承全套申请材料，完成正式申请，跟进审核进度，配合完成核验。',
                        'materials': [
                            f'《{platform}数字遗产法定继承申请表》（按官方要求填写，完整填写逝者与全体继承人信息、账号信息、继承资格说明、遗产分配方案、合规承诺）',
                            '所有材料按官方要求命名、排序、打包',
                            '提交成功凭证与受理编号'
                        ],
                        'templates': [
                            f'{platform}数字遗产法定继承申请表',
                            '继承申请材料补正说明模板'
                        ],
                        'instructions': [
                            '规范提交材料：所有材料按官方要求命名、排序、打包，通过官方指定通道提交，留存提交成功凭证与受理编号',
                            '进度跟进：按官方公示的审核周期定期查询进度，保持预留联系方式畅通，收到补充材料通知后，在时限内完成补正',
                            '核验配合：按官方指引完成全体继承人的身份核验，配合完成继承关系、账号归属的补充确认'
                        ]
                    },
                    {
                        'step': 4,
                        'title': '审核结果处置与兜底维权',
                        'description': '根据审核结果完成资产合规处置，或启动法定继承兜底维权流程。',
                        'materials': [],
                        'templates': [
                            
                        ],
                        'outcome_handling': {
                            'approved': [
                                f'按官方指引与全体继承人确认的分配方案，完成资产合规处置：资金类资产提取划转、虚拟财产合规分配、可继承内容合规导出',
                                f'按全体继承人共识完成账号后续处置，全程留存处置完成凭证，避免继承人间产生纠纷'
                            ],
                            'rejected': [
                                '第一时间确认驳回原因、补正要求、复核通道，补充完善继承资格相关材料后申请复核',
                                '多次复核仍被驳回的，整理全套申请材料、继承关系证明、沟通记录，委托专业继承纠纷律师，通过人民调解、向法院提起继承纠纷诉讼等方式维权，凭生效法律文书向平台申请执行'
                            ]
                        }
                    }
                ]

    return steps


@app.route('/download-inheritance-template/<path:template_name>')
def download_inheritance_template(template_name):
    """下载继承指引模板文件"""
    import os
    from flask import send_file

    # 模板文件映射（只保留文件夹中存在的文件）
    template_map = {
        '数字遗产明细清单': '数字遗产明细清单.xlsx',
        '自书数字遗嘱参考模板': '自书数字遗嘱参考模板.pdf',
        '数字遗产继承公证办理流程说明': '《数字遗产继承公证办理流程说明》.pdf',
        '接受遗赠声明': '《接受遗赠声明》.pdf'
    }

    # 获取对应的文件名
    filename = template_map.get(template_name)

    if not filename:
        flash('未找到该模板文件', 'error')
        return redirect(url_for('inheritance_guide'))

    # 构建模板文件的完整路径
    template_dir = os.path.join(os.path.dirname(__file__), '继承指引模板')
    file_path = os.path.join(template_dir, filename)

    # 检查文件是否存在
    if not os.path.exists(file_path):
        flash('模板文件不存在', 'error')
        return redirect(url_for('inheritance_guide'))

    # 发送文件进行下载
    try:
        return send_file(file_path, as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'下载失败: {str(e)}', 'error')
        return redirect(url_for('inheritance_guide'))


# 路由：情法相伴（故事墙 + 公益普法问答）
@app.route('/stories')
def stories():
    """情法相伴 - 故事墙和公益普法问答"""
    stories = Story.query.filter_by(status='approved').order_by(Story.created_at.desc()).all()
    faqs = FAQ.query.order_by(FAQ.category, FAQ.order).all()
    faq_categories = list(set([faq.category for faq in faqs]))
    faq_categories.sort()
    return render_template('care/index.html', stories=stories, faqs=faqs, faq_categories=faq_categories)


@app.route('/stories/submit', methods=['POST'])
@login_required
def submit_story():
    """提交故事"""
    try:
        title = request.form.get('title')
        category = request.form.get('category')
        content = request.form.get('content')
        author = request.form.get('author')

        if not all([title, category, content, author]):
            return jsonify({'success': False, 'message': '请填写所有必填字段'}), 400

        # 创建故事，状态设为待审核
        story = Story(
            title=title,
            category=category,
            content=content,
            author=author,
            status='pending'
        )

        db.session.add(story)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '故事提交成功！管理员将在审核后发布。'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'提交失败: {str(e)}'}), 500

# 路由：FAQ
@app.route('/faq')
def faq():
    """常见问题解答"""
    search_query = request.args.get('search', '').strip()
    
    if search_query:
        # 搜索功能
        faqs = FAQ.query.filter(
            db.or_(
                FAQ.question.contains(search_query),
                FAQ.answer.contains(search_query)
            )
        ).order_by(FAQ.category, FAQ.order).all()
    else:
        faqs = FAQ.query.order_by(FAQ.category, FAQ.order).all()
    
    # 按指定顺序排序分类
    category_order = ['概念与价值', '安全与管理', '法律与继承', '伦理与隐私']
    # 只保留在category_order中存在的分类，并按指定顺序排序
    categories = [cat for cat in category_order if cat in set(faq.category for faq in faqs)]
    
    return render_template('faq/index.html', faqs=faqs, categories=categories, search_query=search_query)

# 路由：情感关怀（整合故事墙和公益普法问答）
@app.route('/care')
def care():
    """情感关怀 - 整合故事墙和公益普法问答"""
    # 获取故事
    stories = Story.query.filter_by(status='approved').order_by(Story.created_at.desc()).all()
    # 获取FAQ
    faqs = FAQ.query.order_by(FAQ.category, FAQ.order).all()
    category_order = ['概念与价值', '安全与管理', '法律与继承', '伦理与隐私']
    faq_categories = sorted(set(faq.category for faq in faqs), key=lambda x: category_order.index(x) if x in category_order else 999)
    return render_template('care/index.html', stories=stories, faqs=faqs, faq_categories=faq_categories)

# 路由：关于我们
@app.route('/about')
def about():
    """关于我们"""
    return render_template('about.html')

# 路由：后台管理
# 初始化字体和数据库
@app.before_request
def initialize_application():
    """初始化字体和数据库"""
    # 只在第一次请求时初始化
    if not hasattr(initialize_application, 'initialized'):
        try:
            # 1. 首先设置字体
            setup_fonts_on_startup()

            # 2. 然后初始化数据库
            with app.app_context():
                # 确保数据目录存在
                db_path = app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
                db_dir = os.path.dirname(db_path)

                if db_dir and not os.path.exists(db_dir):
                    os.makedirs(db_dir, exist_ok=True)
                    print(f"Created database directory: {db_dir}")

                # 创建所有表
                db.create_all()
                print("Database tables created successfully")

                # 初始化示例数据
                initialize_sample_data()
                print("Sample data initialized successfully")

            initialize_application.initialized = True
        except Exception as e:
            print(f"Application initialization error: {e}")
            import traceback
            traceback.print_exc()
            # 即使初始化失败，也标记为已初始化，避免重复尝试
            initialize_application.initialized = True
def initialize_sample_data():
    """初始化示例数据"""
    # 检查是否已有平台政策数据
    if PlatformPolicy.query.count() > 0:
        # 检查是否有缺失的FAQ，如果有则添加
        existing_questions = set(faq.question for faq in FAQ.query.all())
        
        # 定义所有FAQ
        all_faqs = []
        
        # 添加缺失的FAQ
        for faq in all_faqs:
            if faq.question not in existing_questions:
                db.session.add(faq)
        
        db.session.commit()
        return
    
    # 如果没有平台政策数据，则初始化所有数据

    # 添加平台政策
    policies = [
        PlatformPolicy(
            platform_name='QQ',
            policy_content='账号本身不可被继承；账号内财产可被继承',
            attitude='有限支持',
            inherit_possibility='低',
            legal_basis='根据《QQ号码使用规则》，QQ账号的所有权归腾讯公司，用户仅获得使用权，且该使用权与初始注册人人身身份紧密绑定，平台未允许通过继承转移使用权。但账号内虚拟财产（Q币、QQ钱包余额等）属于用户合法财产，可依法继承。',
            customer_service='综合热线：4006-700-700（9:00-22:00）；在线联系：腾讯客服公众号/小程序、网站kf.qq.com；本地备用：0755-83765566',
            risk_warning='账号本身不可继承，仅账号内虚拟财产可依法继承；需提供完整法律证明文件，审核周期较长'
        ),
        PlatformPolicy(
            platform_name='微信',
            policy_content='账号本身不可被继承；账号内财产可被继承',
            attitude='有限支持',
            inherit_possibility='低',
            legal_basis='根据《微信软件许可及服务协议》，用户仅拥有使用权，不拥有所有权。但微信钱包余额、理财通等财产属于用户合法财产，可依法继承。',
            customer_service='客服热线：95017；在线联系：微信APP内「我-设置-帮助与反馈」、微信/QQ端「腾讯客服」小程序',
            risk_warning='账号本身不可继承，仅账号内虚拟财产可依法继承；账户长期不使用会被冻结，聊天记录可能永久丢失'
        ),
        PlatformPolicy(
            platform_name='抖音',
            policy_content='账号本身不可被继承；账号内财产可被继承；逝者个人信息可依法复制/下载/转移',
            attitude='有限支持',
            inherit_possibility='低',
            legal_basis='根据《抖音用户服务协议》，账号所有权归北京微播视界科技有限公司所有，用户仅获使用权。但抖音钱包余额、抖币等虚拟财产可依法继承。依据《个人信息保护法》及抖音隐私政策3.5条，逝者近亲属可申请复制、下载逝者账号内个人信息。',
            customer_service='客服热线：95152；在线联系：抖音APP内「我-≡-我的客服」、微信/QQ端搜索"抖音"公众号；官方邮箱：feedback@douyin.com',
            risk_warning='账号本身不可继承，逝者未注销账号将自动转为纪念账号；账号连续超过二个月未登录可能被平台收回'
        )
    ]

    for policy in policies:
        db.session.add(policy)

    # 添加FAQ
    # ============================================================
    # 如何添加新的FAQ：
    # 1. 在下面的列表中添加新的 FAQ() 对象
    # 2. 填写 question（问题）、answer（答案）、category（分类）、order（排序）
    # 3. category 可以是：'基础概念'、'保护措施'、'法律问题' 或新增分类
    # 4. order 用于同一分类内的排序，数字越小越靠前
    # ============================================================
    faqs = [
        FAQ(
            question='数字遗产包括哪些内容？',
            answer='数字遗产包括但不限于：社交媒体账号（微信、QQ、抖音等）、电子邮箱、云存储文件、虚拟货币、游戏账号、在线支付账户、博客文章、个人网站、数字相册、音视频文件等。',
            category='基础概念',
            order=1
        ),
        FAQ(
            question='什么是数字遗嘱？',
            answer='数字遗嘱是指用户在生前制定的关于其数字资产如何处理的书面文件，包括账户信息、密码、处理方式等内容的详细说明。它可以指导继承人如何处理您的数字资产，避免账户丢失或数据永久消失。',
            category='基础概念',
            order=2
        ),
        FAQ(
            question='为什么需要规划数字遗产？',
            answer='1. 避免重要数据永久丢失；2. 保护隐私和个人信息；3. 确保资产（如虚拟货币）不被浪费；4. 减轻家人的心理负担；5. 让数字记忆得以传承；6. 避免平台账户被自动删除。',
            category='基础概念',
            order=3
        ),
        FAQ(
            question='数字资产的价值如何评估？',
            answer='数字资产的价值包括：经济价值（虚拟货币、游戏装备、付费内容等）、情感价值（照片、视频、聊天记录等）、实用价值（付费软件、云存储空间等）。建议定期整理和评估您的数字资产。',
            category='基础概念',
            order=4
        ),
        FAQ(
            question='如何保护我的数字遗产？',
            answer='1. 定期备份重要数据到本地或云端；2. 使用密码管理器安全存储密码；3. 创建数字遗嘱并定期更新；4. 告知家人重要账户信息；5. 了解各平台的继承政策；6. 启用双重认证；7. 定期检查账户安全设置。',
            category='保护措施',
            order=1
        ),
        FAQ(
            question='密码安全应该注意什么？',
            answer='1. 使用强密码（大小写字母、数字、特殊符号组合，至少12位）；2. 不要重复使用密码；3. 定期更换密码（每3-6个月）；4. 启用两步验证；5. 使用密码管理器（如LastPass、1Password）；6. 不要在公共场所输入密码；7. 警惕钓鱼网站。',
            category='保护措施',
            order=2
        ),
        FAQ(
            question='如何选择密码管理器？',
            answer='选择密码管理器时考虑：1. 安全性（是否使用AES-256加密）；2. 跨平台支持（手机、电脑、浏览器）；3. 云同步功能；4. 价格（免费版功能）；5. 用户评价和口碑。推荐工具：LastPass、1Password、Bitwarden、KeePass等。',
            category='保护措施',
            order=3
        ),
        FAQ(
            question='数字遗嘱有法律效力吗？',
            answer='数字遗嘱在我国法律体系中尚未明确认定，但可以作为表达意愿的重要依据。根据《民法典》，遗嘱可以采用多种形式，包括打印、录音录像等。数字遗嘱如果能证明是本人真实意愿，可能被参考。建议配合传统遗嘱使用，并咨询专业律师。',
            category='法律问题',
            order=6
        ),
        FAQ(
            question='继承人的法律权利是什么？',
            answer='根据《民法典》，继承人有权继承被继承人的合法财产。但对于数字财产，法律界定尚不明确：1. 虚拟货币（比特币等）通常被视为财产，可以继承；2. 社交媒体账号（微信、QQ等）通常被视为使用权，不可继承；3. 游戏账号和虚拟道具的继承权取决于平台政策；4. 云存储文件可以继承，但需要密码或法律证明。',
            category='法律问题',
            order=7
        ),
        FAQ(
            question='如何证明数字资产的所有权？',
            answer='证明数字资产所有权需要：1. 账户注册信息和登录记录；2. 交易记录和支付凭证（如购买虚拟货币的记录）；3. 电子邮件或聊天记录证明使用情况；4. 平台开具的资产证明；5. 公证处的公证文件；6. 银行流水证明充值记录。建议保留所有相关凭证。',
            category='法律问题',
            order=8
        ),
        FAQ(
            question='如果平台拒绝继承怎么办？',
            answer='1. 查阅平台服务协议，了解具体条款；2. 准备完整的法律文件（死亡证明、继承公证书等）；3. 联系平台客服，说明情况；4. 寻求法律援助，向法院提起诉讼；5. 向消费者协会投诉；6. 通过媒体曝光引起关注。注意：不同平台处理方式不同，需要具体情况具体分析。',
            category='法律问题',
            order=4
        ),
        FAQ(
            question='跨境数字资产继承有什么特殊问题？',
            answer='1. 法律适用问题：涉及不同国家的法律；2. 语言障碍：需要翻译文件；3. 货币兑换：虚拟货币可能需要兑换；4. 税务问题：可能涉及遗产税；5. 时效问题：各国继承时效不同；6. 证据认证：需要领事认证或海牙认证。建议咨询专业国际律师。',
            category='法律问题',
            order=5
        ),
        FAQ(
            question='社交账号能否被作为遗产继承？',
            answer='不能。社交账号具有强烈的人身依附性，其使用权基于用户与平台的合同关系，不属于《民法典》第1122条规定的“遗产”范畴。但如今司法实践中有账号本身不能继承，但账号内的数据内容可以继承的判决倾向。\n法条依据：《民法典》第 1122条：“遗产是自然人死亡时遗留的个人合法财产。依照法律规定或者根据其性质不得继承的遗产，不得继承。”',
            category='法律问题',
            order=1
        ),
        FAQ(
            question='游戏装备、虚拟货币等数字财产如何继承？',
            answer='可以继承，但是受平台用户协议限制。虚拟财产的继承需要平台配合，而平台常以“保护账号安全”为由拒绝。实务中游戏账号的充值余额、已购买未使用消耗的虚拟道具等可以明确估值的财产往往被认为可以继承，而游戏中的成就或其他有关人身性的财产被认定为不可继承。\n法条依据：《民法典》第127条：“法律对数据、网络虚拟财产的保护有规定的，依照其规定。”确认虚拟财产受法律保护；第1122条将其纳入遗产范围。',
            category='法律问题',
            order=2
        ),
        FAQ(
            question='电子邮件、聊天记录等内容是否能被继承人查阅？',
            answer='针对这类隐私性较强的财产应当区分对待。涉及到财产的邮件等应该能够依法被查阅（如银行账单、合同等）涉及到被继承人个人隐私的聊天记录与邮件内容则不能被查阅，除非被继承人同意。\n法条依据：《个人信息保护法》第49条：“自然人死亡的，其近亲属为了自身的合法、正当利益，可以对死者的相关个人信息行使本章规定的查阅、复制、更正、删除等权利；死者生前另有安排的除外。”',
            category='法律问题',
            order=3
        )
    ]

    for faq in faqs:
        db.session.add(faq)

    # 添加示例故事
    stories = [
        Story(
            title='父亲的微信账号',
            content='父亲去世后，我尝试找回他的微信账号，却遇到了重重困难。这让我意识到数字遗产规划的重要性...',
            author='匿名用户',
            category='情感故事',
            status='approved'
        ),
        Story(
            title='数字时代的告别',
            content='在这个数字时代，我们的记忆、情感都存储在云端。如何让这些珍贵的数字遗产得以延续？',
            author='编辑部',
            category='哲思文章',
            status='approved'
        )
    ]

    for story in stories:
        db.session.add(story)

    db.session.commit()

# 错误处理
@app.errorhandler(404)
def not_found(error):
    """404错误处理"""
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    """500错误处理"""
    db.session.rollback()
    return render_template('errors/500.html'), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
