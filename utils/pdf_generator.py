from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
from datetime import datetime

from utils.fonts import register_chinese_font, get_chinese_font_name, get_chinese_bold_font_name

fonts_registered = register_chinese_font()

CHECKED_CHAR = '\u2611'
UNCHECKED_CHAR = '\u2610'

_NO_LINE_START = set('，。、；：！？）》」』】～—…·．')

def _font_supports_char(font_name, char):
    try:
        font = pdfmetrics.getFont(font_name)
        if hasattr(font, 'face') and hasattr(font.face, 'charToGlyph'):
            return ord(char) in font.face.charToGlyph
    except Exception:
        pass
    return False

def _check_marks_strategy():
    fn = get_chinese_font_name()
    if _font_supports_char(fn, CHECKED_CHAR) and _font_supports_char(fn, UNCHECKED_CHAR):
        return 'native', fn
    try:
        pdfmetrics.getFont('UniFont')
        return 'fallback', 'UniFont'
    except Exception:
        pass
    return 'text', None

MARK_STRATEGY, MARK_FONT = _check_marks_strategy()


# 三类资产的所有选项定义（与前端表单一致）
ASSET_CATEGORIES = {
    'social': {
        'title': '一、社交媒体\n（微信、QQ等）',
        'all_actions': ['设为纪念账户', '注销账户', '移交内容给继承人', '其他'],
        'risk_tips': '微信/QQ：账号不可过户，凭证明可申请处理财产或注销。\n通用风险：聊天记录等隐私内容可能无法继承。',
    },
    'cloud': {
        'title': '二、电子邮箱与云存储\n（iCloud、网盘等）',
        'all_actions': ['允许继承人访问', '永久删除/关闭', '其他'],
        'risk_tips': 'iCloud：访问权限极难获取，通常仅支持注销。务必本地备份重要数据。\n网盘服务：多数不支持继承，重要文件需提前下载。',
    },
    'finance': {
        'title': '三、数字货币与金融App\n（比特币、游戏账号等）',
        'all_actions': ['移交私钥/密码', '兑换为法币后继承', '其他'],
        'risk_tips': '比特币：私钥/助记词一旦丢失即永久丧失。\n中心化平台：有遗产继承流程，但需完备法律文件，耗时长。',
    },
}


def _mark(is_checked):
    if MARK_STRATEGY == 'native':
        return CHECKED_CHAR if is_checked else UNCHECKED_CHAR
    elif MARK_STRATEGY == 'fallback':
        mark = CHECKED_CHAR if is_checked else UNCHECKED_CHAR
        return f'<font name="{MARK_FONT}">{mark}</font>'
    else:
        return '[√]' if is_checked else '[ ]'


def _format_actions(category_key, selected_actions):
    cat = ASSET_CATEGORIES.get(category_key, {})
    all_actions = cat.get('all_actions', [])
    lines = []
    for action in all_actions:
        mark = _mark(action in selected_actions)
        lines.append(f'{mark} {action}')
    for sa in selected_actions:
        if sa.startswith('其他：') or sa.startswith('其他:'):
            detail = sa.split('：', 1)[-1].split(':', 1)[-1].strip()
            if detail:
                mark = _mark(True)
                lines[-1] = f'{mark} 其他：{detail}'
    return lines


def generate_will_pdf(will):
    """生成数字资产继承意愿声明书PDF

    Args:
        will: DigitalWill对象

    Returns:
        PDF文件路径
    """
    from flask import current_app
    output_dir = os.path.join(current_app.config.get('DATA_DIR', 'temp_pdfs'), 'temp_pdfs')
    os.makedirs(output_dir, exist_ok=True)

    filename = f'will_{will.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    filepath = os.path.join(output_dir, filename)

    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        rightMargin=48, leftMargin=48, topMargin=45, bottomMargin=45
    )

    styles = getSampleStyleSheet()
    normal_font = get_chinese_font_name()
    bold_font = get_chinese_bold_font_name()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
        fontName=bold_font, fontSize=18, spaceAfter=14, alignment=TA_CENTER, leading=24,
        wordWrap='CJK')

    body_style = ParagraphStyle('CustomBody', parent=styles['BodyText'],
        fontName=normal_font, fontSize=10, spaceAfter=8, alignment=TA_JUSTIFY, leading=17,
        wordWrap='CJK')

    small_style = ParagraphStyle('SmallBody', parent=styles['BodyText'],
        fontName=normal_font, fontSize=9, spaceAfter=4, alignment=TA_JUSTIFY, leading=13,
        wordWrap='CJK')

    story = []

    # 标题
    story.append(Paragraph("数字资产继承意愿声明书", title_style))
    story.append(Spacer(1, 0.1 * inch))

    # 声明人信息
    if will.assets_data:
        declarant_name = will.assets_data.get('declarant_name', '_____________')
        declarant_id = will.assets_data.get('declarant_id', '_____________')
        declarant_nation = will.assets_data.get('declarant_nation', '_____________')
    else:
        declarant_name = declarant_id = declarant_nation = '_____________'

    declarant_data = [['声明人：' + declarant_name, '身份证号：' + declarant_id, '民族：' + declarant_nation]]
    declarant_table = Table(declarant_data, colWidths=[2.0*inch, 2.5*inch, 2.0*inch])
    declarant_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), normal_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10.5),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
    ]))
    story.append(declarant_table)
    story.append(Spacer(1, 0.1 * inch))

    # 核心生效前提
    story.append(Paragraph(
        "本人明确授权，在身故或丧失自主行为能力后，本人指定的联系人（继承人）可依据本声明书所载意愿，对以下各类数字资产进行处置：",
        body_style
    ))
    story.append(Spacer(1, 0.1 * inch))

    # 数字资产类型与处置意愿 - 三列表格（匹配模板PDF格式）
    # 表头
    asset_table_data = [
        [Paragraph('资产类别与示例', ParagraphStyle('th1', parent=styles['BodyText'],
            fontName=bold_font, fontSize=10, textColor=colors.whitesmoke, alignment=TA_CENTER, leading=14, wordWrap='CJK')),
         Paragraph('处置意愿<br/>（请勾选）', ParagraphStyle('th2', parent=styles['BodyText'],
            fontName=bold_font, fontSize=10, textColor=colors.whitesmoke, alignment=TA_CENTER, leading=14, wordWrap='CJK')),
         Paragraph('关键平台政策与风险提示<br/>（供参考）', ParagraphStyle('th3', parent=styles['BodyText'],
            fontName=bold_font, fontSize=10, textColor=colors.whitesmoke, alignment=TA_CENTER, leading=14, wordWrap='CJK'))]
    ]

    cell_style = ParagraphStyle('CellBody', parent=styles['BodyText'],
        fontName=normal_font, fontSize=9.5, alignment=TA_LEFT, leading=15, spaceAfter=1,
        wordWrap='CJK')
    cell_style_center = ParagraphStyle('CellCenter', parent=cell_style, alignment=TA_CENTER)

    for cat_key, cat_info in ASSET_CATEGORIES.items():
        selected = []
        if will.assets_data:
            cat_data = will.assets_data.get(cat_key)
            if cat_data and isinstance(cat_data, dict):
                selected = cat_data.get('actions', [])
            elif cat_data and isinstance(cat_data, list):
                selected = cat_data

        action_lines = _format_actions(cat_key, selected)
        actions_para = Paragraph('<br/>'.join(action_lines), cell_style)

        risk_tips_raw = cat_info.get('risk_tips', '')
        risk_lines = risk_tips_raw.replace('\n', '<br/>')
        risk_para = Paragraph(risk_lines, cell_style)

        title_raw = cat_info['title'].replace('\n', '<br/>')
        title_para = Paragraph(title_raw, cell_style_center)

        asset_table_data.append([title_para, actions_para, risk_para])

    asset_table = Table(asset_table_data, colWidths=[1.4*inch, 2.1*inch, 2.7*inch], repeatRows=1)
    asset_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f0f3f5')),
        ('BACKGROUND', (1, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ced4da')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#2c3e50')),
    ]))
    story.append(asset_table)
    story.append(Spacer(1, 0.2 * inch))

    # 执行人指定
    executor_data = [['本人特此指定以下人员为本声明之执行人：']]
    executor_table = Table(executor_data, colWidths=[6*inch])
    executor_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), normal_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10.5),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(executor_table)

    if will.assets_data:
        primary_name = will.assets_data.get('primary_name', '_____________')
        primary_relation = will.assets_data.get('primary_relation', '_____________')
        primary_phone = will.assets_data.get('primary_phone', '_____________')
        backup_name = will.assets_data.get('backup_name', '_____________')
        backup_relation = will.assets_data.get('backup_relation', '_____________')
        backup_phone = will.assets_data.get('backup_phone', '_____________')
    else:
        primary_name = primary_relation = primary_phone = '_____________'
        backup_name = backup_relation = backup_phone = '_____________'

    story.append(Paragraph(
        f"<b>主要联系人</b>【姓名：{primary_name}，关系：{primary_relation}，电话：{primary_phone}】",
        body_style
    ))
    story.append(Spacer(1, 0.05 * inch))
    story.append(Paragraph(
        f"<b>备用联系人</b>【姓名：{backup_name}，关系：{backup_relation}，电话：{backup_phone}】",
        body_style
    ))
    story.append(Spacer(1, 0.05 * inch))
    story.append(Paragraph(
        "主要联系人将全权负责启动并主导本声明所涉的资产处置流程。仅当主要联系人无法履行职责时，备用联系方可接替其职责。此指定仅为操作授权，并不涉及或改变任何法定继承人之实质财产继承权。",
        small_style
    ))
    story.append(Spacer(1, 0.15 * inch))

    # 法律效力声明
    story.append(Paragraph(
        "本人确认，本声明书旨在清晰表达本人关于数字资产处置的最终意愿，并为继承人提供明确指引。本人理解，本声明书本身并非具有直接强制执行力的法律文件，数字资产的最终归属与处理须遵守《中华人民共和国民法典》等法律规定及各平台合约。本人建议，可将本声明书进行公证，或作为本人正式遗嘱之附件，以增强其法律参考效力。",
        body_style
    ))
    story.append(Spacer(1, 0.2 * inch))

    # 签署栏
    sign_data = [
        ['声明人（签名/指印）：________________________', ''],
        ['', '年      月      日']
    ]
    sign_table = Table(sign_data, colWidths=[4*inch, 2.5*inch])
    sign_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), normal_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10.5),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 1), (1, 1), 'RIGHT'),
    ]))
    story.append(sign_table)

    try:
        doc.build(story)
        if os.path.exists(filepath):
            return filepath
        raise Exception("PDF file was not created")
    except Exception as e:
        print(f"PDF generation error: {e}")
        import traceback
        traceback.print_exc()
        raise


def generate_asset_list_xlsx(user, assets):
    """生成用户数字资产清单Excel文件（匹配模板格式）

    Args:
        user: User对象
        assets: DigitalAsset查询对象或列表

    Returns:
        xlsx文件路径
    """
    from flask import current_app
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from utils.encryption import decrypt_data

    output_dir = os.path.join(current_app.config.get('DATA_DIR', 'temp_pdfs'), 'temp_pdfs')
    os.makedirs(output_dir, exist_ok=True)

    filename = f'assets_{user.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '数字资产明细清单'

    # 样式定义
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    cat_font = Font(name='微软雅黑', size=10, bold=True)
    normal_font_xl = Font(name='微软雅黑', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # 标题行
    ws.merge_cells('A1:E1')
    ws['A1'] = '数字资产明细清单'
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = f'姓名：{user.username}　　邮箱：{user.email}　　生成日期：{datetime.now().strftime("%Y年%m月%d日")}'
    ws['A2'].font = Font(name='微软雅黑', size=10)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    headers = ['资产类别', '平台/应用', '账号', '密码', '备注']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25

    # 按类别分组
    category_labels = {
        'social_media': '社交媒体',
        'email': '电子邮箱',
        'cloud_storage': '云存储与数字内容',
        'virtual_currency': '虚拟资产与数字货币',
        'other': '其他数字资产',
    }

    categories = {}
    for asset in assets:
        cat = asset.category or 'other'
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(asset)

    row_idx = 5
    for cat_key, cat_assets in categories.items():
        cat_label = category_labels.get(cat_key, cat_key)
        cat_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')

        for asset in cat_assets:
            # 尝试解密密码，失败则留空
            password = asset.encrypted_password or ''
            display_pwd = ''
            if password:
                try:
                    display_pwd = decrypt_data(password)
                except Exception:
                    display_pwd = ''  # 解密失败留空

            ws.cell(row=row_idx, column=1, value=cat_label).font = normal_font_xl
            ws.cell(row=row_idx, column=2, value=asset.platform_name or '').font = normal_font_xl
            ws.cell(row=row_idx, column=3, value=asset.account or '').font = normal_font_xl
            ws.cell(row=row_idx, column=4, value=display_pwd).font = normal_font_xl
            ws.cell(row=row_idx, column=5, value=asset.notes or '').font = normal_font_xl

            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = thin_border
                ws.cell(row=row_idx, column=col).alignment = left_align

            row_idx += 1

    if not categories:
        ws.merge_cells(f'A{row_idx}:E{row_idx}')
        ws.cell(row=row_idx, column=1, value='暂无数字资产记录').font = normal_font_xl

    row_idx += 2
    ws.merge_cells(f'A{row_idx}:E{row_idx}')
    ws.cell(row=row_idx, column=1, value='说明：密码栏为空表示该资产密码经加密存储，无法在此导出。如需查看密码，请登录平台在资产详情中查看。').font = Font(name='微软雅黑', size=9, color='666666')
    ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True)

    wb.save(filepath)
    return filepath
